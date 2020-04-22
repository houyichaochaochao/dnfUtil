﻿now_version="2.0.3"
ver_time='200308'

## 未经许可，请勿复制，修改和分发代码。##

import requests
from bs4 import BeautifulSoup
import urllib.request
from urllib import parse
from json import loads
import tkinter.ttk
import tkinter.font
import tkinter.messagebox
from tkinter import *
import openpyxl
from openpyxl import load_workbook
import itertools
import threading
import time
import numpy as np
from collections import Counter
from math import floor
import webbrowser
import tkinter as tk


#https://dunfaoff.com/DawnClass.df

def _from_rgb(rgb):
    """translates an rgb tuple of int to a tkinter friendly color code
    """
    return "#%02x%02x%02x" % rgb

dark_main=_from_rgb((32, 34, 37))
dark_sub=_from_rgb((46, 49, 52))
dark_blue=_from_rgb((29, 30, 36))

    
apikey = "TQS79U4MT11jswCLHq7G260XzXU0JhGC" ##보안코드

exit_calc=0
save_name_list=[]
save_select=0
count_num=0
count_all=0
show_number=0
all_list_num=0


self = tkinter.Tk()
self.title("一键史诗搭配计算器")
self.geometry("710x720+0+0")
self.resizable(False, False)
self.configure(bg=dark_main)
self.iconbitmap(r'ext_img/icon.ico')
guide_font=tkinter.font.Font(family="맑은 고딕", size=10, weight='bold')
mid_font=tkinter.font.Font(family="맑은 고딕", size=14, weight='bold')
big_font=tkinter.font.Font(family="맑은 고딕", size=18, weight='bold')

bg_img=PhotoImage(file = "ext_img/bg_img.png")
bg_wall=tkinter.Label(self,image=bg_img)
bg_wall.place(x=0,y=0)


# 这块是为了读取xlsx里面的内容
load_excel1=load_workbook("DATA.xlsx", data_only=True)

db_one=load_excel1["one"]
opt_one={} # 这里只存储各个武器的属性值 +  名字 + id
name_one={} #  存储各个武器的属性值 +  名字 + id
a=1
for row in db_one.rows:
    row_value=[]
    row_value_cut=[]
    for cell in row:
        row_value.append(cell.value)
        # row_value_cut = row_value[2:] 这里我修改了一下，改到前面去了，不然是无意义的循环
    row_value_cut = row_value[2:]
    opt_one[db_one.cell(a,1).value]=row_value_cut
    name_one[db_one.cell(a,1).value]=row_value
    a=a+1

db_job=load_excel1["lvl"]
opt_job={} # 角色的第三个以后的值
opt_job_ele={} # 前几个值
u=1
for row in db_job.rows:
    row_value=[]
    for cell in row:
        row_value.append(cell.value)
    opt_job[db_job.cell(u,1).value]=row_value[3:]
    opt_job_ele[db_job.cell(u,1).value]=row_value[:3]
    u=u+1
del opt_job["empty"]
del opt_job["직업명"] # 这里是为了删除 最后不用的属性值
jobs=list(opt_job.keys())
    
    
load_preset0=load_workbook("preset.xlsx", data_only=True)
db_custom=load_preset0["custom"]
db_save=load_preset0["one"]
save_name_list=[] # 存储的名字
for i in range(1,11):
    save_name_list.append(db_custom.cell(i,5).value)


########## 首次打开软件预设更新 ###########
try:
    print("DATABASE 版本号 = "+str(db_custom['K1'].value))
    print("客户端 版本号 = "+now_version)
    if str(db_custom['K1'].value) != now_version:
        print("DB 更新版本号")
        db_custom['K1'] = now_version
    if db_custom['H1'].value==None:
        db_custom['G1']="up_stat"
        db_custom['H1']=0
    if db_custom['H2'].value==None:
        db_custom['G2']="bless_style"
        db_custom['H2']=3
    if db_custom['H3'].value==None:
        db_custom['G3']="crux_style"
        db_custom['H3']=2
    if db_custom['H4'].value==None:
        db_custom['G4']="bless_plt"
        db_custom['H4']=2
    if db_custom['H5'].value==None:
        db_custom['G5']="bless_cri"
        db_custom['H5']=1
    if db_custom['H6'].value==None:
        db_custom['G6']="up_stat_b"
        db_custom['H6']=0
        
    if db_custom['B14'].value==None:
        db_custom['A14']="ele_inchant"
        db_custom['B14']=116
    if db_custom['B15'].value==None:
        db_custom['A15']="ele_ora"
        db_custom['B15']=20
    if db_custom['B16'].value==None:
        db_custom['A16']="ele_gem"
        db_custom['B16']=7
    if db_custom['B17'].value==None:
        db_custom['A17']="ele_skill"
    db_custom['B17']=0  ## 자속강 비활성화 禁用磁通钢
    if db_custom['B18'].value==None:
        db_custom['A18']="ele_mob_resist"
        db_custom['B18']=50
    if db_custom['B19'].value==None:
        db_custom['A19']="ele_buf_anti"
        db_custom['B19']=60
    if db_save['A257'].value==None:
        db_save['A257']='13390150';db_save['B257']='+5 完美控制'
        db_save['C257']=0;db_save['D257']=0;db_save['E257']=0;db_save['F257']=0;db_save['G257']=0
        db_save['H257']=0;db_save['I257']=0;db_save['J257']=0;db_save['K257']=0;db_save['L257']=0
    if db_save['A258'].value==None:
        db_save['A258']='22390240';db_save['B258']='+4 先知的项链'
        db_save['C258']=0;db_save['D258']=0;db_save['E258']=0;db_save['F258']=0;db_save['G258']=0
        db_save['H258']=0;db_save['I258']=0;db_save['J258']=0;db_save['K258']=0;db_save['L258']=0
    if db_save['A259'].value==None:
        db_save['A259']='21400340';db_save['B259']='+4 독을 머금은 가시장갑'
        db_save['C259']=0;db_save['D259']=0;db_save['E259']=0;db_save['F259']=0;db_save['G259']=0
        db_save['H259']=0;db_save['I259']=0;db_save['J259']=0;db_save['K259']=0;db_save['L259']=0
    if db_save['A260'].value==None:
        db_save['A260']='23390450';db_save['B260']='+5 할기의 링'
        db_save['C260']=0;db_save['D260']=0;db_save['E260']=0;db_save['F260']=0;db_save['G260']=0
        db_save['H260']=0;db_save['I260']=0;db_save['J260']=0;db_save['K260']=0;db_save['L260']=0
    if db_save['A261'].value==None:
        db_save['A261']='31400540';db_save['B261']='+4 청면수라의 가면'
        db_save['C261']=0;db_save['D261']=0;db_save['E261']=0;db_save['F261']=0;db_save['G261']=0
        db_save['H261']=0;db_save['I261']=0;db_save['J261']=0;db_save['K261']=0;db_save['L261']=0
    if db_save['A262'].value==None:
        db_save['A262']='32410650';db_save['B262']='+5 적귀의 차원석'
        db_save['C262']=0;db_save['D262']=0;db_save['E262']=0;db_save['F262']=0;db_save['G262']=0
        db_save['H262']=0;db_save['I262']=0;db_save['J262']=0;db_save['K262']=0;db_save['L262']=0
    if db_save['A263'].value==None:
        db_save['A263']='33390750';db_save['B263']='+5 패스트퓨처 이어링'
        db_save['C263']=0;db_save['D263']=0;db_save['E263']=0;db_save['F263']=0;db_save['G263']=0
        db_save['H263']=0;db_save['I263']=0;db_save['J263']=0;db_save['K263']=0;db_save['L263']=0
        


    load_preset0.save("preset.xlsx") # 这里是保存一下
    
except PermissionError as error:
    tkinter.messagebox.showerror("错误","更新失败. 请重新运行.")
load_excel1.close()
load_preset0.close()

    
## 计算函数##
def calc():
    global result_window
    try:
        result_window.destroy()
    except NameError as error:
        pass
    if select_perfect.get() == '慢速':
        set_perfect=1
    else:
        set_perfect=0
    showsta(text="正在准备组合算法驱动...")
    start_time = time.time()
    load_excel=load_workbook("DATA.xlsx",data_only=True)

    db_one=load_excel["one"]
    opt_one={}
    name_one={}
    a=1
    for row in db_one.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_one[db_one.cell(a,1).value]=row_value_cut # 这个是基础属性值 key = 武器名字   value = 武器属性
        name_one[db_one.cell(a,1).value]=row_value # 这个是包括id和名字 key = 武器名字  value = 所有值一行内的，和我想的一样，后面根本用不到
        a=a+1

    # 这里装入的是装备的各种组合，已经写好了，分为2,3,5 词条属性分别为1,2,3
    b=1        
    db_set=load_excel["set"]
    opt_set={}
    for row in db_set.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_set[db_set.cell(b,1).value]=row_value_cut ## DB 装入 ## 所以这里只要cut了
        b=b+1

    # 这里是考虑各个装备的buff 效果 根据id识别
    c=1        
    db_buf=load_excel["buf"]
    opt_buf={}
    name_buf={}
    for row in db_buf.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_buf[db_buf.cell(c,1).value]=row_value_cut ## DB 装入 ##
        name_buf[db_buf.cell(c,1).value]=row_value
        c=c+1

    # 这里是buf的效果，但是没看懂
    d=1        
    db_buflvl=load_excel["buflvl"]
    opt_buflvl={}
    for row in db_buflvl.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = [0] + row_value[1:]
        opt_buflvl[db_buflvl.cell(d,1).value]=row_value_cut
        d=d+1

    # 这里加载一些值
    load_presetc=load_workbook("preset.xlsx", data_only=True)
    db_preset=load_presetc["custom"]
    try:
        ele_skill=int(opt_job_ele[jobup_select.get()][1])  # 这里是选择职业
        ele_in = (int(db_preset["B14"].value) + int(db_preset["B15"].value) + int(db_preset["B16"].value) +
                  int(ele_skill) - int(db_preset["B18"].value) + int(db_preset["B19"].value) + 13)
    except KeyError:
        tkinter.messagebox.showwarning(title='警告', message='请先选择职业')


    global count_num, count_all, show_number,all_list_num, max_setopt
    count_num=0;count_all=0;show_number=0


    if jobup_select.get()[-4:] == "(奶系)":
        active_eff_one=15
        active_eff_set=18-3
    else:
        active_eff_one=21
        active_eff_set=27-3

    if time_select.get() == "60秒(觉醒占比↓)":
        lvl_shift=6
    else:
        lvl_shift=0

    job_lv1=opt_job[jobup_select.get()][11+lvl_shift]
    job_lv2=opt_job[jobup_select.get()][12+lvl_shift]
    job_lv3=opt_job[jobup_select.get()][13+lvl_shift]
    job_lv4=opt_job[jobup_select.get()][14+lvl_shift]
    job_lv5=opt_job[jobup_select.get()][15+lvl_shift]
    job_lv6=opt_job[jobup_select.get()][16+lvl_shift]
    job_pas0=opt_job[jobup_select.get()][0]
    job_pas1=opt_job[jobup_select.get()][1]
    job_pas2=opt_job[jobup_select.get()][2]
    job_pas3=opt_job[jobup_select.get()][3]
    
    
    extra_dam=0
    extra_cri=0
    extra_bon=0
    extra_all=0
    extra_pas2=0
    extra_sta=0
    extra_att=0
    if style_select.get() == '伟大的意志':
        extra_cri=extra_cri+18
        extra_sta=extra_sta+4
    if style_select.get() == '使徒降临':
        extra_bon=extra_bon+12
        extra_sta=extra_sta+3
    if style_select.get() == '国庆称号':
        extra_cri=extra_cri+15 
    if creature_select.get() == '普通宠物':
        extra_bon=extra_bon+12
        extra_sta=extra_sta+10
    if creature_select.get() == '至尊宠物':
        extra_bon=extra_bon+15
        extra_sta=extra_sta+12

    if req_cool.get()=='X(纯伤害)':
        cool_on=0
    else:
        cool_on=1
    list11=[];list12=[];list13=[];list14=[];list15=[] # 这里对应上衣 下装 头肩 腰带 鞋子
    list21=[];list22=[];list23=[];list31=[];list32=[];list33=[] # 这里分别对应 手镯 项链  戒指  辅助装备 魔法石  耳环
    list_setnum=[];list_godnum=[]
    for i in range(1010,1999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list11.append('1'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(2010,2999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list12.append('1'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(3010,3999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list13.append('1'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(4010,4999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list14.append('1'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(5010,5999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list15.append('1'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(1010,1999):
        try:
            if eval('select_item["tg2{}"]'.format(i)) == 1:
                list21.append('2'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(2010,2999):
        try:
            if eval('select_item["tg2{}"]'.format(i)) == 1:
                list22.append('2'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(3010,3999):
        try:
            if eval('select_item["tg2{}"]'.format(i)) == 1:
                list23.append('2'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(1010,1999):
        try:
            if eval('select_item["tg3{}"]'.format(i)) == 1:
                list31.append('3'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(2010,2999):
        try:
            if eval('select_item["tg3{}"]'.format(i)) == 1:
                list32.append('3'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    for i in range(3010,3999):
        try:
            if eval('select_item["tg3{}"]'.format(i)) == 1:
                list33.append('3'+str(i))
                list_setnum.append(str(i)[1:])
        except KeyError as error:
            c=1
    algo_list=['11','12','13','14','15','21','22','23','31','32','33']  #
    if select_perfect.get() == '快速':
        for i in list_setnum:
            if list_setnum.count(i)==1:
                if i[-1]!='1':
                    for ca in algo_list:
                        try:
                            eval("list{}.remove('{}{}')".format(ca,ca,i))
                        except ValueError as error:
                            c=1

    for know_one in know_list:
        if eval('select_item["tg{}"]'.format(know_one)) == 1:
            eval('list{}.append(str({}))'.format(know_one[0:2],know_one))

    if len(list11)==0:
        list11.append('11360');list12.append('12360');list13.append('13360');list14.append('14360');list15.append('15360')
    elif len(list12)==0:
        list11.append('11360');list12.append('12360');list13.append('13360');list14.append('14360');list15.append('15360')
    elif len(list13)==0:
        list11.append('11360');list12.append('12360');list13.append('13360');list14.append('14360');list15.append('15360')
    elif len(list14)==0:
        list11.append('11360');list12.append('12360');list13.append('13360');list14.append('14360');list15.append('15360')
    elif len(list15)==0:
        list11.append('11360');list12.append('12360');list13.append('13360');list14.append('14360');list15.append('15360')
        
    if len(list21)==0 and len(list22)==0:
        list21.append('21370');list22.append('22370');list23.append('23370')
    elif len(list22)==0 and len(list23)==0:
        list21.append('21370');list22.append('22370');list23.append('23370')
    elif len(list23)==0 and len(list21)==0:
        list21.append('21370');list22.append('22370');list23.append('23370')
        
    if len(list31)==0 and len(list32)==0:
        list31.append('31380');list32.append('32380');list33.append('33380')
    elif len(list32)==0 and len(list33)==0:
        list31.append('31380');list32.append('32380');list33.append('33380')
    elif len(list33)==0 and len(list31)==0:
        list31.append('31380');list32.append('32380');list33.append('33380')

    if len(list21)==0:
        list21.append('21370')
    if len(list22)==0:
        list22.append('22370')
    if len(list23)==0:
        list23.append('23370')
    if len(list31)==0:
        list31.append('31380')
    if len(list32)==0:
        list32.append('32380')
    if len(list33)==0:
        list33.append('33380')

    #print(list11)
    #print(list12)
    #print(list13)
    #print(list14)
    #print(list15)
    #print(list21)
    #print(list22)
    #print(list23)
    #print(list31)
    #print(list32)
    #print(list33)
        
    for i in range(0,75):
        if wep_select.get() == wep_list[i]:
            wep_num=(str(i+111001),)
            
    items=[list11,list12,list13,list14,list15,list21,list22,list23,list31,list32,list33]
    all_list_num=len(list11)*len(list12)*len(list13)*len(list14)*len(list15)*len(list21)*len(list22)*len(list23)*len(list31)*len(list32)*len(list33)
    if all_list_num > 500000000:
        tkinter.messagebox.showerror('提示',"组合超过5亿种.\n无法进行.\n请关闭一定史诗")
        showsta(text='已终止')
        return
    elif all_list_num > 100000000:
        ask_msg2=tkinter.messagebox.askquestion('确认',"情况的数量超过1亿种.\可能需要30分钟以上.\n确定进行吗？")
        if ask_msg2 == 'no':
            showsta(text='已终止')
            return
    elif all_list_num > 30000000:
        ask_msg2=tkinter.messagebox.askquestion('确认',"情况的数量超过3000种.\n可能需要很长时间.\n确定进行吗？")
        if ask_msg2 == 'no':
            showsta(text='已终止')
            return
    if set_perfect ==1 and all_list_num > 3000000:
        tkinter.messagebox.showerror('提示',"功能不支持很多情况。")
        showsta(text='已终止')
        return

    try:
        all_list=list(itertools.product(*items))
        showsta(text='开始计算')
    except MemoryError as error:
        tkinter.messagebox.showerror('内存误差',"已超过可用内存.")
        showsta(text='已中止')
        return
    global exit_calc
    if jobup_select.get()[4:7] != "奶爸" and jobup_select.get()[4:7] != "奶妈" and jobup_select.get()[4:7] != "奶萝":
    
        # 代码名称
        # 0调节器 1추공  2증 3크 4추 5속추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 액티브 / 16~19 패시브 /20 신화여부/21세트코드/22쿨감보정/23 원쿨감
        # 단품:24~33 저장소 / 34 二觉캐액티브
        getone=opt_one.get
        save_list={}
        max_setopt=0
        show_number=1
        for calc_now in all_list:
            if exit_calc==1:
                showsta(text='已终止')
                return
            god=0
            for o in calc_now:
                god=god+int(o[-1])
            if god < 2:
                set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)] 
                set_val=Counter(set_list)
                del set_val['136','137','138']
                setopt_num=sum([floor(x*0.7) for x in set_val.values()])+god
                if setopt_num >= max_setopt-set_perfect :
                    set_on=[];setapp=set_on.append
                    setcount=set_list.count
                    set_oncount=set_on.count
                    onecount=calc_now.count
                    for i in range(101,136):
                        if setcount(str(i))==2:
                            setapp(str(i)+"1")
                        if 4>=setcount(str(i))>=3:
                            setapp(str(i)+"2")
                        if setcount(str(i))==5:
                            setapp(str(i)+"3")
                    for i in range(136,139):
                        if setcount(str(i))==2:
                            setapp(str(i)+"0")
                        if 4>=setcount(str(i))>=3:
                            setapp(str(i)+"1")
                        if setcount(str(i))==5:
                            setapp(str(i)+"2")
                    if onecount('32410650')==1:
                        if onecount('21400340')==1:
                            setapp('1401')
                        elif onecount('31400540')==1:
                            setapp('1401')
                    if max_setopt <= setopt_num-god*set_perfect:
                        max_setopt = setopt_num-god*set_perfect
                    try:
                        calc_wep=wep_num+calc_now
                    except UnboundLocalError:
                        tk.messagebox.showwarning(title="警告",message="请先选择武器名字")
                        return
                    else:
                        pass
                    damage=0
                    base_array=np.array([0,0,extra_dam,extra_cri,extra_bon,0,extra_all,extra_att,extra_sta,ele_in,0,1,0,0,0,0,0,0,extra_pas2,0,0,0,0,0,0,0,0,0])
                    
                    skiper=0
                    try:
                        for_calc=tuple(set_on)+calc_wep
                        oneone = len(for_calc)
                        oneonelist = []
                        for i in range(oneone):
                            no_cut = getone(for_calc[i])  ## 11번 스증
                            cut = np.array(no_cut[0:20] + no_cut[22:23] + no_cut[34:35] + no_cut[38:44])
                            skiper = (skiper / 100 + 1) * (cut[11] / 100 + 1) * 100 - 100
                            oneonelist.append(cut)
                        for i in range(oneone):
                            base_array = base_array + oneonelist[i]
                        # 코드 이름
                        # 0추스탯 1추공 2증 3크 4추 5속추
                        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
                        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 2각캐특수액티브 /22~27 액티브레벨링
                    except UnboundLocalError:
                        pass #  这里是不是也是return好一些

                    if set_oncount('1201')==1 and onecount('32200')==1:
                        base_array[3]=base_array[3]-5
                    if onecount('33230')==1 or onecount('33231')==1:
                        if onecount('31230')==0:
                            base_array[4]=base_array[4]-10
                        if onecount('32230')==0:
                            base_array[9]=base_array[9]-40
                    if onecount('15340')==1 or onecount('23340')==1 or onecount('33340')==1 or onecount('33341')==1:
                        if set_oncount('1341')==0 and set_oncount('1342') ==0:
                            if onecount('15340')==1:
                                base_array[9]=base_array[9]-20
                            elif onecount('23340')==1:
                                base_array[2]=base_array[2]-10
                            elif onecount('33340')==1:
                                base_array[6]=base_array[6]-5
                            else:
                                base_array[9]=base_array[9]-4
                                base_array[2]=base_array[2]-2
                                base_array[6]=base_array[6]-1
                                base_array[8]=base_array[8]-1.93
                    if onecount('11111')==1:
                        if set_oncount('1112')==1 or set_oncount('1113')==1:
                            base_array[20]=base_array[20]+10
                    if onecount('11301')==1:
                        if onecount('22300')!=1:
                            base_array[4]=base_array[4]-10
                            base_array[7]=base_array[7]+10
                        if onecount('31300')!=1:
                            base_array[4]=base_array[4]-10
                            base_array[7]=base_array[7]+10
                    
                    base_array[11]=skiper
                    real_bon=base_array[4]+base_array[5]*(base_array[9]*0.0045+1.05)
                    actlvl=((base_array[active_eff_one]+base_array[22]*job_lv1+base_array[23]*job_lv2+base_array[24]*job_lv3+
                            base_array[25]*job_lv4+base_array[26]*job_lv5+base_array[27]*job_lv6)/100+1)
                    paslvl=((100+base_array[16]*job_pas0)/100)*((100+base_array[17]*job_pas1)/100)*((100+base_array[18]*job_pas2)/100)*((100+base_array[19]*job_pas3)/100)
                    damage=((base_array[2]/100+1)*(base_array[3]/100+1)*(real_bon/100+1)*(base_array[6]/100+1)*(base_array[7]/100+1)*
                            (base_array[8]/100+1)*(base_array[9]*0.0045+1.05)*(base_array[10]/100+1)*(skiper/100+1)*(base_array[12]/100+1)*
                            actlvl*paslvl*((54500+3.31*base_array[0])/54500)*((4800+base_array[1])/4800)*(1+cool_on*base_array[20]/100)/(1.05+0.0045*int(ele_skill)))
                    
                    base_array[4]=real_bon
                    try:
                        save_list[damage]=[calc_wep,base_array]
                    except UnboundLocalError:
                        pass
                    count_num=count_num+1
                else:
                    count_all=count_all+1
            else:
                count_all=count_all+1
        show_number=0
        showsta(text='结果统计中')
        
        ranking=[]
        for j in range(0,5):
            try:
                for i in save_list.keys():
                    if max(list(save_list.keys()))==i:
                        ranking.append((i,save_list.get(i)))
                        del save_list[i]
                        showsta(text='结果统计中'+str(j)+" / 5")
            except RuntimeError as error:
                passss=1
                
        show_result(ranking,'deal',ele_skill)

    else:
        base_b=10+int(db_preset['H2'].value)+int(db_preset['H4'].value)+int(db_preset['H5'].value)+1
        base_c=12+int(db_preset['H3'].value)+1
        base_pas0=0
        base_pas0_c=3
        base_pas0_b=0
        base_stat_s=4166+74-126+int(db_preset['H1'].value)
        base_stat_d=int(db_preset['H6'].value)-int(db_preset['H1'].value)
        base_stat_h=4308-45-83+int(db_preset['H1'].value)
        base_pas0_1=0
        load_presetc.close()
        lvlget=opt_buflvl.get
        #코드 이름
        #0 체정 1 지능
        #祝福 2 스탯% 3 물공% 4 마공% 5 독공%
        #아포 6 고정 7 스탯%
        #8 축렙 9 포렙
        #10 아리아/보징증폭
        #11 전직패 12 보징 13 각패1 14 각패2 15 二觉 16 각패3
        #17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
        save_list1={}
        save_list2={}
        save_list3={}
        setget=opt_buf.get
        max_setopt=0
        show_number=1
        for calc_now in all_list:
            if exit_calc==1:
                showsta(text='已终止')
                return
            base_array=np.array([base_stat_h,base_stat_s,0,0,0,0,0,0,base_b,base_c,0,base_pas0,base_pas0_1,0,0,0,0,0,0,0,0])
            try:
                calc_wep=wep_num+calc_now
            except UnboundLocalError:
                tk.messagebox.showwarning("警告","必须先选择武器")
                return
            god=0
            for o in calc_now:
                god=god+int(o[-1])

            if god < 2:
                set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)]
                set_val=Counter(set_list)
                del set_val['136','137','138']
                setopt_num=sum([floor(x*0.7) for x in set_val.values()])+god
                if setopt_num >= max_setopt-set_perfect :
                    set_on=[];setapp=set_on.append
                    setcount=set_list.count
                    for i in range(101,136):
                        if setcount(str(i))==2:
                            setapp(str(i)+"1")
                        if 4>=setcount(str(i))>=3:
                            setapp(str(i)+"2")
                        if setcount(str(i))==5:
                            setapp(str(i)+"3")
                    for i in range(136,139):
                        if setcount(str(i))==2:
                            setapp(str(i)+"0")
                        if 4>=setcount(str(i))>=3:
                            setapp(str(i)+"1")
                        if setcount(str(i))==5:
                            setapp(str(i)+"2")
                    if max_setopt <= setopt_num-god*set_perfect:
                        max_setopt = setopt_num-god*set_perfect
                    b_stat=10.24 ##탈리스만
                    b_phy=0
                    b_mag=0
                    b_ind=0
                    c_per=0
                    for_calc=tuple(set_on)+calc_wep
                    oneone=len(for_calc)
                    oneonelist=[]
                    for i in range(oneone):
                        no_cut=np.array(setget(for_calc[i]))             ## 2 3 4 5 7
                        try:
                            base_array=base_array+no_cut  # 提醒需要换奶装
                        except:
                            tk.messagebox.showwarning(title="警告",message="必须先选择奶爸武器")
                            return
                        b_stat=(b_stat/100+1)*(no_cut[2]/100+1)*100-100
                        b_phy=(b_phy/100+1)*(no_cut[3]/100+1)*100-100
                        b_mag=(b_mag/100+1)*(no_cut[4]/100+1)*100-100
                        b_ind=(b_ind/100+1)*(no_cut[5]/100+1)*100-100
                        c_per=(c_per/100+1)*(no_cut[7]/100+1)*100-100
                        oneonelist.append(no_cut)
                        
                    #0 체정 1 지능
                    #祝福 2 스탯% 3 물공% 4 마공% 5 독공%
                    #아포 6 고정 7 스탯%
                    #8 축렙 9 포렙
                    #10 아리아/보징증폭
                    #11 전직패 12 보징 13 각패1 14 각패2 15 二觉 16 각패3
                    #17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
                    if jobup_select.get()[4:7] == "奶爸":
                        b_base_att=lvlget('hol_b_atta')[int(base_array[8])]
                        stat_pas0lvl_b=lvlget('pas0')[int(base_array[11])+base_pas0_b]+lvlget('hol_pas0_1')[int(base_array[12])]
                        stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+base_pas0_c]+lvlget('hol_pas0_1')[int(base_array[12])]
                        stat_pas1lvl=lvlget('hol_pas1')[int(base_array[13])]
                        stat_pas2lvl=lvlget('hol_act2')[int(base_array[15])]
                        stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
                        stat_b=base_array[0]+stat_pas0lvl_b+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]+base_stat_d
                        stat_c=base_array[0]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]
                        b_stat_calc=int(int(lvlget('hol_b_stat')[int(base_array[8])]*(b_stat/100+1))*(stat_b/630+1))
                        b_phy_calc=int(int(b_base_att*(b_phy/100+1))*(stat_b/630+1))
                        b_mag_calc=int(int(b_base_att*(b_mag/100+1))*(stat_b/630+1))
                        b_ind_calc=int(int(b_base_att*(b_ind/100+1))*(stat_b/630+1))
                        b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
                        c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
                        pas1_calc=int(lvlget('hol_pas1_out')[int(base_array[13])]+213+base_array[17])
                        pas1_out=str(int(lvlget('hol_pas1_out')[int(base_array[13])]+213+base_array[17]))+"  ("+str(int(20+base_array[13]))+"렙)"
                        save1=str(b_stat_calc)+"/"+str(b_average)+"   ["+str(int(stat_b))+"("+str(int(base_array[8]))+"렙)]"

                    else:
                        if jobup_select.get()[4:7] == "奶妈":
                            b_value=675
                            aria=1.25+0.05*base_array[10]
                        if jobup_select.get()[4:7] == "奶萝":
                            b_value=665
                            aria=(1.20+0.05*base_array[10])*1.20
                            
                        b_base_att=lvlget('se_b_atta')[int(base_array[8])]
                        stat_pas0lvl_b=lvlget('pas0')[int(base_array[11])+int(base_pas0_b)]
                        stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+int(base_pas0_c)]
                        stat_pas1lvl=lvlget('se_pas1')[int(base_array[13])]+base_array[18]
                        stat_pas2lvl=lvlget('se_pas2')[int(base_array[14])]
                        stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
                        stat_b=base_array[1]+stat_pas0lvl_b+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+base_stat_d
                        stat_c=base_array[1]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl
                        b_stat_calc=int(int(lvlget('se_b_stat')[int(base_array[8])]*(b_stat/100+1))*(stat_b/b_value+1)*aria)
                        b_phy_calc=int(int(b_base_att*(b_phy/100+1)*(stat_b/b_value+1))*aria)
                        b_mag_calc=int(int(b_base_att*(b_mag/100+1)*(stat_b/b_value+1))*aria)
                        b_ind_calc=int(int(b_base_att*(b_ind/100+1)*(stat_b/b_value+1))*aria)
                        b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
                        c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
                        pas1_calc=int(stat_pas1lvl+442)
                        pas1_out=str(int(stat_pas1lvl+442))+"  ("+str(int(20+base_array[13]))+"렙)"
                        save1=str(b_stat_calc)+"("+str(int(b_stat_calc/aria))+")/ "+str(b_average)+"("+str(int(b_average/aria))+")\n                  ["+str(int(stat_b))+"("+str(int(base_array[8]))+"렙)]"

                    save2=str(c_calc)+"    ["+str(int(stat_c))+"("+str(int(base_array[9]))+"렙)]"
                    ##1축 2포 3합
                    save_list1[((15000+b_stat_calc)/250+1)*(2650+b_average)]=[calc_wep,[save1,save2,pas1_out]]
                    save_list2[((15000+c_calc)/250+1)*2650]=[calc_wep,[save1,save2,pas1_out]]
                    save_list3[((15000+pas1_calc+c_calc+b_stat_calc)/250+1)*(2650+b_average)]=[calc_wep,[save1,save2,pas1_out]]
                                            
                    count_num=count_num+1
                else:
                    count_all=count_all+1
            else:
                count_all=count_all+1
        show_number=0
        showsta(text='结果统计中')

        ranking1=[];ranking2=[];ranking3=[]
        for j in range(0,5):
            try:
                for i in save_list1.keys():
                    if max(list(save_list1.keys()))==i:
                        ranking1.append((i,save_list1.get(i)))
                        del save_list1[i]
            except RuntimeError as error:
                passss=1

        for j in range(0,5):
            try:
                for i in save_list2.keys():
                    if max(list(save_list2.keys()))==i:
                        ranking2.append((i,save_list2.get(i)))
                        del save_list2[i]
            except RuntimeError as error:
                passss=1

        for j in range(0,5):
            try:
                for i in save_list3.keys():
                    if max(list(save_list3.keys()))==i:
                        ranking3.append((i,save_list3.get(i)))
                        del save_list3[i]
            except RuntimeError as error:
                passss=1
        ranking=[ranking1,ranking2,ranking3]
        show_result(ranking,'buf',ele_skill)
    load_excel.close()
    showsta(text='输出完成')
    print("时间 = "+str(time.time() - start_time)+"秒")
    
def calc_thread():
    threading.Thread(target=calc,daemon=True).start()

def show_result(rank_list,job_type,ele_skill):
    global result_window
    result_window=tkinter.Toplevel(self)
    result_window.attributes("-topmost", True)
    result_window.geometry("585x402+800+400")
    result_window.title("计算结果")
    result_window.resizable(False,False)
    global canvas_res
    canvas_res = Canvas(result_window, width=587, height=404, bd=0)
    canvas_res.place(x=-2,y=-2)
    if job_type=='deal':
        result_bg=tkinter.PhotoImage(file='ext_img/bg_result.png')
    else:
        result_bg=tkinter.PhotoImage(file='ext_img/bg_result2.png')
    canvas_res.create_image(293,202,image=result_bg)

    global image_list
    global res_img11,res_img12,res_img13,res_img14,res_img15,res_img21,res_img22,res_img23,res_img31,res_img32,res_img33,wep_select,jobup_select

    if job_type=='deal': #########这个deal啥意思还是没明白

        global result_image_on,rank_dam,rank_stat,rank_stat2,req_cool,res_dam,res_stat,res_stat2
        rank_dam=[0,0,0,0,0]
        rank_setting=[0,0,0,0,0]
        rss=[0,0,0,0,0]
        result_image_on=[{},{},{},{},{}]
        try:
            rank_dam[0]=str(int(100*rank_list[0][0]))+"%"
            rank_setting[0]=rank_list[0][1][0] ##0号是排名
            rss[0]=rank_list[0][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[0][str(i)]=image_list[j]
            rank_dam[1]=str(int(100*rank_list[1][0]))+"%"
            rank_setting[1]=rank_list[1][1][0] 
            rss[1]=rank_list[1][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[1][str(i)]=image_list[j]
            rank_dam[2]=str(int(100*rank_list[2][0]))+"%"
            rank_setting[2]=rank_list[2][1][0] 
            rss[2]=rank_list[2][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[2][str(i)]=image_list[j]
            rank_dam[3]=str(int(100*rank_list[3][0]))+"%"
            rank_setting[3]=rank_list[3][1][0] 
            rss[3]=rank_list[3][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[3][str(i)]=image_list[j]
            rank_dam[4]=str(int(100*rank_list[4][0]))+"%"
            rank_setting[4]=rank_list[4][1][0] 
            rss[4]=rank_list[4][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[4][str(i)]=image_list[j]
        except IndexError as error:
            c=1

        # 0추스탯 1추공 2증 3크 4추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 二觉캐특수액티브 /22~27 액티브레벨링
        rank_stat=[0,0,0,0,0]
        rank_stat2=[0,0,0,0,0]
        for i in range(0,5):
            try:
                rank_stat[i]=("增伤= "+str(int(rss[i][2]))+
                              "%\n爆伤= "+str(int(rss[i][3]))+
                              "%\n白字= "+str(int(rss[i][4]))+
                              "%\n所攻= "+str(int(rss[i][6]))+
                              "%\n三攻= "+str(int(rss[i][7]))+
                              "%\n力智= "+str(int(rss[i][8]))+
                              "%\n属强= "+str(int(rss[i][9]))+
                              "\n持续= "+str(int(rss[i][10]))+
                              "%\n技攻= "+str(int(rss[i][11]))+
                              "%\n特殊= "+str(int(rss[i][12]))+
                              "%\n\n速度= "+str(int(rss[i][13]))+
                              "%\n暴击= "+str(int(rss[i][14]))+"%")
                rank_stat2[i]=("   <主动>\n 1~45级= "+str(round(rss[i][22],1))+
                               "렙\n  50等级= "+str(int(rss[i][23]))+
                               "렙\n60~80等级= "+str(round(rss[i][24],1))+
                               "렙\n  85等级= "+str(int(rss[i][25]))+
                               "렙\n  95等级= "+str(int(rss[i][26]))+
                               "렙\n 100等级= "+str(int(rss[i][27]))+
                               "렙\n\n   <被动>\n 전직패= "+str(round(rss[i][16],1))+
                               "렙\n  一觉= "+str(int(rss[i][17]))+
                               "렙\n  二觉= "+str(int(rss[i][18]))+
                               "렙\n  三觉= "+str(int(rss[i][19]))+"렙")
            except TypeError as error:
                c=1

        cool_check=req_cool.get()[2:6]
        canvas_res.create_text(122,145,text=cool_check,font=guide_font,fill='white')
        if int(ele_skill) != 0:
            canvas_res.create_text(122,170,text="技能属强补正="+str(int(ele_skill))+" / 역보정%="+str(round(100*(1.05/(1.05+int(ele_skill)*0.0045)-1),1))+"%",font=guide_font,fill='white')
        res_dam=canvas_res.create_text(122,125,text=rank_dam[0],font=mid_font,fill='white')
        res_stat=canvas_res.create_text(65,293,text=rank_stat[0],fill='white')
        res_stat2=canvas_res.create_text(185,293,text=rank_stat2[0],fill='white')

        res_img11=canvas_res.create_image(57,57,image=result_image_on[0]['11'])
        res_img12=canvas_res.create_image(27,87,image=result_image_on[0]['12'])
        res_img13=canvas_res.create_image(27,57,image=result_image_on[0]['13'])
        res_img14=canvas_res.create_image(57,87,image=result_image_on[0]['14'])
        res_img15=canvas_res.create_image(27,117,image=result_image_on[0]['15'])
        res_img21=canvas_res.create_image(189,57,image=result_image_on[0]['21'])
        res_img22=canvas_res.create_image(219,57,image=result_image_on[0]['22'])
        res_img23=canvas_res.create_image(219,87,image=result_image_on[0]['23'])
        res_img31=canvas_res.create_image(189,87,image=result_image_on[0]['31'])
        res_img32=canvas_res.create_image(219,117,image=result_image_on[0]['32'])
        res_img33=canvas_res.create_image(189,117,image=result_image_on[0]['33'])
        cn1=0
        for j in range(0,5):
            try:
                for i in [11,12,13,14,15,21,22,23,31,32,33]:
                    canvas_res.create_image(268+cn1*29,67+78*j,image=result_image_on[j][str(i)])
                    cn1=cn1+1
                cn1=0
                canvas_res.create_text(346,34+78*j,text=rank_dam[j],font=mid_font,fill='white')
            except KeyError as error:
                c=1
        length=len(rank_list)

    elif job_type=='buf': ########################## 这个是算的奶装
        load_presetr=load_workbook("preset.xlsx", data_only=True)
        r_preset=load_presetr["custom"]
        global result_image_on1,result_image_on2,result_image_on3,rank_buf1,rank_buf2,rank_buf3, rank_type_buf, res_buf, res_img_list, res_buf_list, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3, res_buf_type_what
        rank_type_buf=3
        rank_setting1=[0,0,0,0,0]
        rank_setting2=[0,0,0,0,0]
        rank_setting3=[0,0,0,0,0]
        result_image_on1=[{},{},{},{},{}]
        result_image_on2=[{},{},{},{},{}]
        result_image_on3=[{},{},{},{},{}]
        rank_buf1=[0,0,0,0,0]
        rank_buf2=[0,0,0,0,0]
        rank_buf3=[0,0,0,0,0]
        rank_buf_ex1=[0,0,0,0,0]
        rank_buf_ex2=[0,0,0,0,0]
        rank_buf_ex3=[0,0,0,0,0]
        ## rank_setting[rank]=rank_list[a][rank][b][c]
        ## a: 0=祝福,1=크오,2=합계
        ## b: 0=계수,1=스펙or증가량
        ## c: b에서 1 선택시, 0=스펙, 1=증가량
        try:
            rank_setting3[0]=rank_list[2][0][1][0]  ##2번째 숫자가 랭킹임
            rank_setting2[0]=rank_list[1][0][1][0]
            rank_setting1[0]=rank_list[0][0][1][0]
            rank_buf3[0]=int(rank_list[2][0][0]/10)
            rank_buf2[0]=int(rank_list[1][0][0]/10)
            rank_buf1[0]=int(rank_list[0][0][0]/10)
            rank_buf_ex3[0]=rank_list[2][0][1][1]
            rank_buf_ex2[0]=rank_list[1][0][1][1]
            rank_buf_ex1[0]=rank_list[0][0][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[0][str(i)]=image_list[j]
                for j in rank_setting2[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[0][str(i)]=image_list[j]
                for j in rank_setting1[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[0][str(i)]=image_list[j] ##
            rank_setting3[1]=rank_list[2][1][1][0]
            rank_setting2[1]=rank_list[1][1][1][0]
            rank_setting1[1]=rank_list[0][1][1][0]
            rank_buf3[1]=int(rank_list[2][1][0]/10)
            rank_buf2[1]=int(rank_list[1][1][0]/10)
            rank_buf1[1]=int(rank_list[0][1][0]/10)
            rank_buf_ex3[1]=rank_list[2][1][1][1]
            rank_buf_ex2[1]=rank_list[1][1][1][1]
            rank_buf_ex1[1]=rank_list[0][1][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[1][str(i)]=image_list[j] 
                for j in rank_setting2[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[1][str(i)]=image_list[j]
                for j in rank_setting1[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[1][str(i)]=image_list[j] ##
            rank_setting3[2]=rank_list[2][2][1][0]
            rank_setting2[2]=rank_list[1][2][1][0]
            rank_setting1[2]=rank_list[0][2][1][0]
            rank_buf3[2]=int(rank_list[2][2][0]/10)
            rank_buf2[2]=int(rank_list[1][2][0]/10)
            rank_buf1[2]=int(rank_list[0][2][0]/10)
            rank_buf_ex3[2]=rank_list[2][2][1][1]
            rank_buf_ex2[2]=rank_list[1][2][1][1]
            rank_buf_ex1[2]=rank_list[0][2][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[2][str(i)]=image_list[j]
                for j in rank_setting2[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[2][str(i)]=image_list[j]
                for j in rank_setting1[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[2][str(i)]=image_list[j] ##
            rank_setting3[3]=rank_list[2][3][1][0]
            rank_setting2[3]=rank_list[1][3][1][0]
            rank_setting1[3]=rank_list[0][3][1][0]
            rank_buf3[3]=int(rank_list[2][3][0]/10)
            rank_buf2[3]=int(rank_list[1][3][0]/10)
            rank_buf1[3]=int(rank_list[0][3][0]/10)
            rank_buf_ex3[3]=rank_list[2][3][1][1]
            rank_buf_ex2[3]=rank_list[1][3][1][1]
            rank_buf_ex1[3]=rank_list[0][3][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[3][str(i)]=image_list[j]
                for j in rank_setting2[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[3][str(i)]=image_list[j]
                for j in rank_setting1[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[3][str(i)]=image_list[j] ##
            rank_setting3[4]=rank_list[2][4][1][0]
            rank_setting2[4]=rank_list[1][4][1][0]
            rank_setting1[4]=rank_list[0][4][1][0]
            rank_buf3[4]=int(rank_list[2][4][0]/10)
            rank_buf2[4]=int(rank_list[1][4][0]/10)
            rank_buf1[4]=int(rank_list[0][4][0]/10)
            rank_buf_ex3[4]=rank_list[2][4][1][1]
            rank_buf_ex2[4]=rank_list[1][4][1][1]
            rank_buf_ex1[4]=rank_list[0][4][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[4][str(i)]=image_list[j]
                for j in rank_setting2[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[4][str(i)]=image_list[j]
                for j in rank_setting1[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[4][str(i)]=image_list[j] ##
        except IndexError as error:
            c=1
            
        canvas_res.create_text(122,193,text="커스텀 祝福+"+str(int(r_preset['H2'].value)+int(r_preset['H4'].value)+int(r_preset['H5'].value))+"렙 / "+"커스텀 一觉+"+str(int(r_preset['H3'].value))+"렙\n祝福스탯+"+str(int(r_preset['H6'].value))+" / 一觉 스탯+"+str(int(r_preset['H1'].value)),font=guide_font,fill='white')

        res_buf=canvas_res.create_text(122,125,text=rank_buf3[0],font=mid_font,fill='white')
        res_buf_type_what=canvas_res.create_text(122,145,text="总体 标准",font=guide_font,fill='white')
        res_buf_ex1=canvas_res.create_text(123,247,text="祝福="+rank_buf_ex3[0][0],font=guide_font,fill='white')
        res_buf_ex2=canvas_res.create_text(123-17,282,text="一觉="+rank_buf_ex3[0][1],font=guide_font,fill='white')
        res_buf_ex3=canvas_res.create_text(123-44,312,text="一觉损失="+rank_buf_ex3[0][2],font=guide_font,fill='white')

        res_img11=canvas_res.create_image(57,57,image=result_image_on3[0]['11'])
        res_img12=canvas_res.create_image(27,87,image=result_image_on3[0]['12'])
        res_img13=canvas_res.create_image(27,57,image=result_image_on3[0]['13'])
        res_img14=canvas_res.create_image(57,87,image=result_image_on3[0]['14'])
        res_img15=canvas_res.create_image(27,117,image=result_image_on3[0]['15'])
        res_img21=canvas_res.create_image(189,57,image=result_image_on3[0]['21'])
        res_img22=canvas_res.create_image(219,57,image=result_image_on3[0]['22'])
        res_img23=canvas_res.create_image(219,87,image=result_image_on3[0]['23'])
        res_img31=canvas_res.create_image(189,87,image=result_image_on3[0]['31'])
        res_img32=canvas_res.create_image(219,117,image=result_image_on3[0]['32'])
        res_img33=canvas_res.create_image(189,117,image=result_image_on3[0]['33'])
        cn1=0
        res_img_list={}
        res_buf_list={}
        for j in range(0,5):
            try:
                for i in [11,12,13,14,15,21,22,23,31,32,33]:
                    temp_res=canvas_res.create_image(268+cn1*29,67+78*j,image=result_image_on3[j][str(i)])
                    res_img_list[str(j)+str(i)]=temp_res
                    cn1=cn1+1
                cn1=0
                temp_buf=canvas_res.create_text(346,34+78*j,text=rank_buf3[j],font=mid_font,fill='white')
                res_buf_list[j]=temp_buf
            except KeyError as error:
                c=1
        length=len(rank_list[0])
        type1_img=tkinter.PhotoImage(file='ext_img/type_bless.png')
        type2_img=tkinter.PhotoImage(file='ext_img/type_crux.png')
        type3_img=tkinter.PhotoImage(file='ext_img/type_all.png')
        rank_type_but1=tkinter.Button(result_window,command=lambda:change_rank_type(1),image=type1_img,bg=dark_main,borderwidth=0,activebackground=dark_main);rank_type_but1.place(x=8,y=337)
        rank_type_but2=tkinter.Button(result_window,command=lambda:change_rank_type(2),image=type2_img,bg=dark_main,borderwidth=0,activebackground=dark_main);rank_type_but2.place(x=84,y=337)
        rank_type_but3=tkinter.Button(result_window,command=lambda:change_rank_type(3),image=type3_img,bg=dark_main,borderwidth=0,activebackground=dark_main);rank_type_but3.place(x=160,y=337)
        rank_type_but1.image=type1_img
        rank_type_but2.image=type2_img
        rank_type_but3.image=type3_img
        
        load_presetr.close()

    wep_name=wep_select.get()
    job_name=jobup_select.get()[:-4]
    job_up_name=jobup_select.get()[-4:]
    canvas_res.create_text(122,20,text=wep_name,font=guide_font,fill='white')
    canvas_res.create_text(122,50,text="<직업>",font=guide_font,fill='white')
    canvas_res.create_text(122,70,text=job_name,font=guide_font,fill='white')
    canvas_res.create_text(122,87,text=job_up_name,font=guide_font,fill='white')
    
    show_detail_img=tkinter.PhotoImage(file='ext_img/show_detail.png')
    
    res_bt1=tkinter.Button(result_window,command=lambda:change_rank(0,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue);res_bt1.place(x=486,y=20+78*0) # 一定会有一个
    res_bt2=tkinter.Button(result_window,command=lambda:change_rank(1,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    res_bt3=tkinter.Button(result_window,command=lambda:change_rank(2,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    res_bt4=tkinter.Button(result_window,command=lambda:change_rank(3,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    res_bt5=tkinter.Button(result_window,command=lambda:change_rank(4,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    if length>1:
        res_bt2.place(x=486,y=20+78*1)
    if length>2:
        res_bt3.place(x=486,y=20+78*2)
    if length>3:
        res_bt4.place(x=486,y=20+78*3)
    if length>4:
        res_bt5.place(x=486,y=20+78*4)

    canvas_res.image=result_bg
    res_bt1.image=show_detail_img

def change_rank(now,job_type):
    global image_list,canvas_res, res_img11,res_img12,res_img13,res_img14,res_img15,res_img21,res_img22,res_img23,res_img31,res_img32,res_img33
    if job_type =='deal':
        global res_dam,res_stat,res_stat2,rank_stat,rank_stat2,result_image_on
        try:
            image_changed=result_image_on[now]
            canvas_res.itemconfig(res_img11,image=image_changed['11']) # 1 开头是装备
            canvas_res.itemconfig(res_img12,image=image_changed['12'])
            canvas_res.itemconfig(res_img13,image=image_changed['13'])
            canvas_res.itemconfig(res_img14,image=image_changed['14'])
            canvas_res.itemconfig(res_img15,image=image_changed['15'])
            canvas_res.itemconfig(res_img21,image=image_changed['21']) # 2 开头是首饰
            canvas_res.itemconfig(res_img22,image=image_changed['22'])
            canvas_res.itemconfig(res_img23,image=image_changed['23'])
            canvas_res.itemconfig(res_img31,image=image_changed['31']) # 3 开头是特殊装备
            canvas_res.itemconfig(res_img32,image=image_changed['32'])
            canvas_res.itemconfig(res_img33,image=image_changed['33'])
            canvas_res.itemconfig(res_dam,text=rank_dam[now])
            canvas_res.itemconfig(res_stat,text=rank_stat[now])
            canvas_res.itemconfig(res_stat2,text=rank_stat2[now])
        except KeyError as error:
            c=1

    elif job_type =='buf':
        global result_image_on1,result_image_on2,result_image_on3,rank_buf1,rank_buf2,rank_buf3, rank_type_buf, res_buf, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3
        try:
            if rank_type_buf==1:
                image_changed=result_image_on1[now]
                rank_changed=rank_buf1[now]
                rank_buf_ex_changed=rank_buf_ex1
            elif rank_type_buf==2:
                image_changed=result_image_on2[now]
                rank_changed=rank_buf2[now]
                rank_buf_ex_changed=rank_buf_ex2
            elif rank_type_buf==3:
                image_changed=result_image_on3[now]
                rank_changed=rank_buf3[now]
                rank_buf_ex_changed=rank_buf_ex3
            canvas_res.itemconfig(res_buf,text=rank_changed)
            canvas_res.itemconfig(res_buf_ex1,text="祝福="+rank_buf_ex_changed[now][0])
            canvas_res.itemconfig(res_buf_ex2,text="一觉="+rank_buf_ex_changed[now][1])
            canvas_res.itemconfig(res_buf_ex3,text="一觉패="+rank_buf_ex_changed[now][2])                
            canvas_res.itemconfig(res_img11,image=image_changed['11'])
            canvas_res.itemconfig(res_img12,image=image_changed['12'])
            canvas_res.itemconfig(res_img13,image=image_changed['13'])
            canvas_res.itemconfig(res_img14,image=image_changed['14'])
            canvas_res.itemconfig(res_img15,image=image_changed['15'])
            canvas_res.itemconfig(res_img21,image=image_changed['21'])
            canvas_res.itemconfig(res_img22,image=image_changed['22'])
            canvas_res.itemconfig(res_img23,image=image_changed['23'])
            canvas_res.itemconfig(res_img31,image=image_changed['31'])
            canvas_res.itemconfig(res_img32,image=image_changed['32'])
            canvas_res.itemconfig(res_img33,image=image_changed['33'])
        except KeyError as error:
            c=1
        

def change_rank_type(in_type):
    global image_list,canvas_res, res_img11,res_img12,res_img13,res_img14,res_img15,res_img21,res_img22,res_img23,res_img31,res_img32,res_img33
    global result_image_on1,result_image_on2,result_image_on3,rank_buf1,rank_buf2,rank_buf3, rank_type_buf, res_img_list, res_buf_list, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3, res_buf_type_what
    if in_type==1:
        rank_type_buf=1
        image_changed=result_image_on1[0]
        image_changed_all=result_image_on1
        rank_changed=rank_buf1
        rank_buf_ex_changed=rank_buf_ex1
        type_changed="祝福 기준"
    elif in_type==2:
        rank_type_buf=2
        image_changed=result_image_on2[0] 
        image_changed_all=result_image_on2
        rank_changed=rank_buf2
        rank_buf_ex_changed=rank_buf_ex2
        type_changed="一觉 기준"
    elif in_type==3:
        rank_type_buf=3
        image_changed=result_image_on3[0]
        image_changed_all=result_image_on3
        rank_changed=rank_buf3
        rank_buf_ex_changed=rank_buf_ex3
        type_changed="총합 기준"
    canvas_res.itemconfig(res_buf_type_what,text=type_changed)
    canvas_res.itemconfig(res_buf_ex1,text="祝福="+rank_buf_ex_changed[0][0])
    canvas_res.itemconfig(res_buf_ex2,text="一觉="+rank_buf_ex_changed[0][1])
    canvas_res.itemconfig(res_buf_ex3,text="一觉패="+rank_buf_ex_changed[0][2])            
    canvas_res.itemconfig(res_buf,text=rank_changed[0])
    canvas_res.itemconfig(res_img11,image=image_changed['11'])
    canvas_res.itemconfig(res_img12,image=image_changed['12'])
    canvas_res.itemconfig(res_img13,image=image_changed['13'])
    canvas_res.itemconfig(res_img14,image=image_changed['14'])
    canvas_res.itemconfig(res_img15,image=image_changed['15'])
    canvas_res.itemconfig(res_img21,image=image_changed['21'])
    canvas_res.itemconfig(res_img22,image=image_changed['22'])
    canvas_res.itemconfig(res_img23,image=image_changed['23'])
    canvas_res.itemconfig(res_img31,image=image_changed['31'])
    canvas_res.itemconfig(res_img32,image=image_changed['32'])
    canvas_res.itemconfig(res_img33,image=image_changed['33'])
    cn2=0
    for j in range(0,5):
            try:
                for i in [11,12,13,14,15,21,22,23,31,32,33]:
                    canvas_res.itemconfig(res_img_list[str(j)+str(i)],image=image_changed_all[j][str(i)])
                    cn2=cn2+2
                cn2=0
                canvas_res.itemconfig(res_buf_list[j],text=rank_changed[j],font=mid_font,fill='white')
            except KeyError as error:
                c=1


def costum():
    global custom_window # 可以对函数外面的变量进行操作，这来自global 的
    custom_window=tkinter.Toplevel(self)
    custom_window.attributes("-topmost", True) # 这里我删掉也没影响为啥？？？
    custom_window.geometry("620x400+750+20")

    load_preset=load_workbook("preset.xlsx",data_only=True)
    db_preset=load_preset["custom"]
    
    tkinter.Label(custom_window,text="<输出环境>",font=mid_font).place(x=100,y=10)
    tkinter.Label(custom_window,text="属性攻击=",font=guide_font).place(x=10,y=50)
    ele_list=['火','冰','光','暗']
    ele_type=tkinter.ttk.Combobox(custom_window,width=5,values=ele_list); ele_type.place(x=80,y=52) ##     
    ele_type.set(db_preset['B1'].value)
    tkinter.Label(custom_window,text="冷却补正比例=          %",font=guide_font).place(x=160,y=50) ##Y11/Z11
    cool_con=tkinter.Entry(custom_window,width=5);cool_con.place(x=230,y=52)
    cool_con.insert(END,db_preset['B2'].value)
    
    tkinter.Label(custom_window,text="<特殊装备补正>",font=mid_font).place(x=100,y=85)
    tkinter.Label(custom_window,text=" 输入窗口的数值会以对应百分比加成最终数值",fg="Red").place(x=30,y=120)
    tkinter.Label(custom_window,text="歧路腰带=          %",font=guide_font).place(x=160,y=155) ##O164
    cus1=tkinter.Entry(custom_window,width=5);cus1.place(x=230,y=157)
    cus1.insert(END,db_preset['B3'].value)
    tkinter.Label(custom_window,text="歧路鞋子=          %",font=guide_font).place(x=160,y=185) ##O180
    cus2=tkinter.Entry(custom_window,width=5);cus2.place(x=230,y=187)
    cus2.insert(END,db_preset['B4'].value)
    tkinter.Label(custom_window,text="经验等级=          ",font=guide_font).place(x=160,y=215) ##G276
    lvl_list=['传说↓','英雄↑']
    cus3=tkinter.ttk.Combobox(custom_window,width=5,values=lvl_list); cus3.place(x=230,y=217)
    cus3.set(db_preset['B12'].value)
    tkinter.Label(custom_window,text="恍惚增幅等级=          강",font=guide_font).place(x=160,y=245)
    lvl_list=[10,11,12,13]
    cus4=tkinter.ttk.Combobox(custom_window,width=2,values=lvl_list); cus4.place(x=230,y=247)
    cus4.set(db_preset['B13'].value)

    
    tkinter.Label(custom_window,text="不息上衣=          %",font=guide_font).place(x=10,y=155) ##O100
    cus6=tkinter.Entry(custom_window,width=5);cus6.place(x=80,y=157)
    cus6.insert(END,db_preset['B5'].value)
    tkinter.Label(custom_window,text="不息裤子=          %",font=guide_font).place(x=10,y=185) ##O127
    cus7=tkinter.Entry(custom_window,width=5);cus7.place(x=80,y=187)
    cus7.insert(END,db_preset['B6'].value)
    tkinter.Label(custom_window,text="不息护肩=          %",font=guide_font).place(x=10,y=215) ##O147
    cus8=tkinter.Entry(custom_window,width=5);cus8.place(x=80,y=217)
    cus8.insert(END,db_preset['B7'].value)
    tkinter.Label(custom_window,text="不息腰带=          %",font=guide_font).place(x=10,y=245) ##O163
    cus9=tkinter.Entry(custom_window,width=5);cus9.place(x=80,y=247)
    cus9.insert(END,db_preset['B8'].value)
    tkinter.Label(custom_window,text="不息鞋子=          %",font=guide_font).place(x=10,y=275) ##O179
    cus10=tkinter.Entry(custom_window,width=5);cus10.place(x=80,y=277)
    cus10.insert(END,db_preset['B9'].value)
    tkinter.Label(custom_window,text="不息2件套=           %",font=guide_font).place(x=10,y=305) ##O295
    cus11=tkinter.Entry(custom_window,width=5);cus11.place(x=80,y=307)
    cus11.insert(END,db_preset['B10'].value)
    tkinter.Label(custom_window,text="不息3件套=           %",font=guide_font).place(x=10,y=335) ##O296,O297
    cus12=tkinter.Entry(custom_window,width=5);cus12.place(x=80,y=337)
    cus12.insert(END,db_preset['B11'].value)

    tkinter.Label(custom_window,text="<奶量增幅相关>",font=mid_font,fg='blue').place(x=410,y=5)
    tkinter.Label(custom_window,text="补正辅助角色的表现",fg="Red").place(x=350,y=33)
    tkinter.Label(custom_window,text="面板体精智+          ",font=guide_font).place(x=320,y=75) ##
    c_stat=tkinter.Entry(custom_window,width=7);c_stat.place(x=390,y=82)
    c_stat.insert(END,db_preset['H1'].value)
    tkinter.Label(custom_window,text="面板体精智+          ",font=guide_font).place(x=470,y=75) ##
    b_stat=tkinter.Entry(custom_window,width=7);b_stat.place(x=540,y=82)
    b_stat.insert(END,db_preset['H6'].value)
    three=[0,1,2,3];two=[0,1,2]
    tkinter.Label(custom_window,text="祝福称号=",font=guide_font).place(x=320,y=110)
    b_style_lvl=tkinter.ttk.Combobox(custom_window,width=5,values=three); b_style_lvl.place(x=390,y=112) ##     
    b_style_lvl.set(db_preset['H2'].value)
    tkinter.Label(custom_window,text="一觉称号=",font=guide_font).place(x=470,y=110)
    c_style_lvl=tkinter.ttk.Combobox(custom_window,width=5,values=two); c_style_lvl.place(x=540,y=112) ##     
    c_style_lvl.set(db_preset['H3'].value)
    tkinter.Label(custom_window,text="祝福等级=",font=guide_font).place(x=320,y=140)
    b_plt=tkinter.ttk.Combobox(custom_window,width=5,values=two); b_plt.place(x=390,y=142) ##
    b_plt.set(db_preset['H4'].value)
    tkinter.Label(custom_window,text="祝福等级=",font=guide_font).place(x=470,y=140)
    b_cri=tkinter.ttk.Combobox(custom_window,width=5,values=[0,1]); b_cri.place(x=540,y=142) ##
    b_cri.set(db_preset['H5'].value)

    tkinter.Label(custom_window,text="<属强提升>",font=mid_font).place(x=410,y=175)
    tkinter.Label(custom_window,text="基础属强=",font=guide_font).place(x=470,y=210)
    ele1=tkinter.Entry(custom_window,width=7); ele1.place(x=540,y=212) ##
    ele1.insert(END,db_preset['B14'].value)
    tkinter.Label(custom_window,text="其他属强=",font=guide_font).place(x=470,y=240)
    ele2=tkinter.Entry(custom_window,width=7); ele2.place(x=540,y=242) ##
    ele2.insert(END,db_preset['B15'].value)
    tkinter.Label(custom_window,text="勋章属强=",font=guide_font).place(x=470,y=270)
    ele3=tkinter.Entry(custom_window,width=7); ele3.place(x=540,y=272) ##
    ele3.insert(END,db_preset['B16'].value)
    tkinter.Label(custom_window,text="技能属强= ",font=guide_font).place(x=320,y=210)
    ele4=tkinter.Entry(custom_window,width=7); ele4.place(x=390,y=212) ## 
    ele4.insert(END,db_preset['B17'].value)
    tkinter.Label(custom_window,text="怪物属抗=",font=guide_font).place(x=320,y=240)
    ele5=tkinter.Entry(custom_window,width=7); ele5.place(x=390,y=242) ##
    ele5.insert(END,db_preset['B18'].value)
    tkinter.Label(custom_window,text="辅助减抗=",font=guide_font).place(x=320,y=270)
    ele6=tkinter.Entry(custom_window,width=7); ele6.place(x=390,y=272) ##
    ele6.insert(END,db_preset['B19'].value)

    load_preset.close()
    save_command=lambda:save_custom(ele_type.get(),cool_con.get(),cus1.get(),cus2.get(),cus3.get(),cus4.get(),
                                    cus6.get(),cus7.get(),cus8.get(),cus9.get(),cus10.get(),cus11.get(),cus12.get(),
                                    c_stat.get(),b_stat.get(),b_style_lvl.get(),c_style_lvl.get(),b_plt.get(),b_cri.get(),
                                    ele1.get(),ele2.get(),ele3.get(),ele4.get(),ele5.get(),ele6.get())
    tkinter.Button(custom_window,text="保存",font=mid_font,command=save_command,bg="lightyellow").place(x=190,y=295)
    
def save_custom(ele_type,cool_con,cus1,cus2,cus3,cus4,cus6,cus7,cus8,cus9,cus10,cus11,cus12,c_stat,b_stat,b_style_lvl,c_style_lvl,b_plt,b_cri,ele1,ele2,ele3,ele4,ele5,ele6):
    try:
        load_excel3=load_workbook("DATA.xlsx")
        load_preset1=load_workbook("preset.xlsx")
        db_custom1=load_preset1["custom"]
        db_save_one=load_excel3["one"]
        db_save_set=load_excel3["set"]
        
        db_custom1['B1']=ele_type
        if ele_type == '화':
            db_save_one['L181']=0;db_save_one['L165']=0;db_save_one['L149']=24;db_save_one['L129']=0
        elif ele_type == '수':
            db_save_one['L181']=0;db_save_one['L165']=24;db_save_one['L149']=0;db_save_one['L129']=0
        elif ele_type == '명':
            db_save_one['L181']=24;db_save_one['L165']=0;db_save_one['L149']=0;db_save_one['L129']=0
        elif ele_type == '암':
            db_save_one['L181']=0;db_save_one['L165']=0;db_save_one['L149']=0;db_save_one['L129']=24
            
        db_custom1['B2']=float(cool_con)
        for i in range(1,257):
            db_save_one.cell(i, 25).value=db_save_one.cell(i, 26).value*float(cool_con)/100
        for i in range(259,351):
            db_save_one.cell(i, 25).value=db_save_one.cell(i, 26).value*float(cool_con)/100
        for i in range(1,93):
            db_save_set.cell(i, 25).value=db_save_one.cell(i, 26).value*float(cool_con)/100
        
        db_custom1['B3']=float(cus1);db_save_one['O164']=float(cus1)
        db_custom1['B4']=float(cus2);db_save_one['O180']=float(cus2)
        db_custom1['B5']=float(cus6);db_save_one['O100']=float(cus6);db_save_one['O101']=float(cus6)
        db_custom1['B6']=float(cus7);db_save_one['O127']=float(cus7)
        db_custom1['B7']=float(cus8);db_save_one['O147']=float(cus8)
        db_custom1['B8']=float(cus9);db_save_one['O163']=float(cus9)
        db_custom1['B9']=float(cus10);db_save_one['O179']=float(cus10)
        db_custom1['B10']=float(cus11);db_save_one['O295']=float(cus11)
        db_custom1['B11']=float(cus12);db_save_one['O296']=float(cus12);db_save_one['O297']=float(cus12)
        db_custom1['B12']=cus3
        if cus3=='전설↓':
            db_save_one['J86']=34;db_save_one['F120']=34;db_save_one['N140']=34;db_save_one['L156']=68;db_save_one['K172']=34;db_save_one['G276']=40;
        else:
            db_save_one['J86']=35;db_save_one['F120']=35;db_save_one['N140']=35;db_save_one['L156']=72;db_save_one['K172']=35;db_save_one['G276']=41;
        db_custom1['B13']=cus4
        db_save_one['N189']=int(cus4)+4;db_save_one['N190']=int(cus4)+4;db_save_one['K205']=int(cus4)+4;db_save_one['E214']=int(cus4)+4

        db_custom1['H1']=c_stat
        db_custom1['H6']=b_stat
        db_custom1['H2']=b_style_lvl
        db_custom1['H3']=c_style_lvl
        db_custom1['H4']=b_plt
        db_custom1['H5']=b_cri

        db_custom1['B14']=ele1
        db_custom1['B15']=ele2
        db_custom1['B16']=ele3
        db_custom1['B17']=ele4
        db_custom1['B18']=ele5
        db_custom1['B19']=ele6
        
        load_preset1.save("preset.xlsx")
        load_preset1.close()
        load_excel3.save("DATA.xlsx")
        load_excel3.close()
        custom_window.destroy()
        tkinter.messagebox.showinfo("通知","保存完成")
    except PermissionError as error:
        tkinter.messagebox.showerror("错误","请重试")


def load_checklist():
    ask_msg1=tkinter.messagebox.askquestion('确认',"找回保存明细？")
    for snum in range(0,10):
        if save_select.get() == save_name_list[snum]:
            ssnum1=snum
    if ask_msg1 == 'yes':
        load_preset3=load_workbook("preset.xlsx")
        db_load_check=load_preset3["one"]
        load_cell=db_load_check.cell
        k=1
        for i in range(1,264):
            if load_cell(i,2+ssnum1).value == 1:
                try:
                    select_item['tg{}'.format(load_cell(i,1).value)]=1
                except KeyError as error:
                    passss=1
            elif load_cell(i,2+ssnum1).value == 0:
                try:
                    select_item['tg{}'.format(load_cell(i,1).value)]=0
                except KeyError as error:
                    passss=1
        load_preset3.close()
        check_equipment()
        for i in range(101,136):
            check_set(i)

def save_checklist():
    ask_msg2=tkinter.messagebox.askquestion('确认',"确认保存吗？")
    for snum in range(0,10):
        if save_select.get() == save_name_list[snum]:
            ssnum2=snum
    try:
        if ask_msg2 == 'yes':
            load_preset4=load_workbook("preset.xlsx")
            db_save_check=load_preset4["one"]
            save_cell=db_save_check.cell
            opt_save={}
            for i in range(1,264):
                opt_save[save_cell(i,1).value]=i

            for code in opt_save.keys():
                try:
                    if eval("select_item['tg{}']".format(code)) == 1:
                        save_cell(opt_save[code],2+ssnum2).value=1
                except KeyError as error:
                    passss1=1
                    
                try:
                    if eval("select_item['tg{}']".format(code)) == 0:
                        save_cell(opt_save[code],2+ssnum2).value=0
                except KeyError as error:
                    passss1=1
                
                passss=1
            load_preset4.save("preset.xlsx")
            load_preset4.close()
            tkinter.messagebox.showinfo("通知","保存完成")
    except PermissionError as error:
        tkinter.messagebox.showerror("错误","请关闭重试")

def change_list_name():
    global change_window
    change_window=tkinter.Toplevel(self)
    change_window.geometry("190x320+750+200")
    tkinter.Label(change_window,text="1套").place(x=20,y=10)
    tkinter.Label(change_window,text="2套").place(x=20,y=35)
    tkinter.Label(change_window,text="3套").place(x=20,y=60)
    tkinter.Label(change_window,text="4套").place(x=20,y=85)
    tkinter.Label(change_window,text="5套").place(x=20,y=110)
    tkinter.Label(change_window,text="6套").place(x=20,y=135)
    tkinter.Label(change_window,text="7套").place(x=20,y=160)
    tkinter.Label(change_window,text="8套").place(x=20,y=185)
    tkinter.Label(change_window,text="9套").place(x=20,y=210)
    tkinter.Label(change_window,text="10套").place(x=20,y=235)
    entry1=tkinter.Entry(change_window,width=10);entry1.place(x=95,y=12);entry1.insert(END,save_name_list[0])
    entry2=tkinter.Entry(change_window,width=10);entry2.place(x=95,y=37);entry2.insert(END,save_name_list[1])
    entry3=tkinter.Entry(change_window,width=10);entry3.place(x=95,y=62);entry3.insert(END,save_name_list[2])
    entry4=tkinter.Entry(change_window,width=10);entry4.place(x=95,y=87);entry4.insert(END,save_name_list[3])
    entry5=tkinter.Entry(change_window,width=10);entry5.place(x=95,y=112);entry5.insert(END,save_name_list[4])
    entry6=tkinter.Entry(change_window,width=10);entry6.place(x=95,y=137);entry6.insert(END,save_name_list[5])
    entry7=tkinter.Entry(change_window,width=10);entry7.place(x=95,y=162);entry7.insert(END,save_name_list[6])
    entry8=tkinter.Entry(change_window,width=10);entry8.place(x=95,y=187);entry8.insert(END,save_name_list[7])
    entry9=tkinter.Entry(change_window,width=10);entry9.place(x=95,y=212);entry9.insert(END,save_name_list[8])
    entry10=tkinter.Entry(change_window,width=10);entry10.place(x=95,y=237);entry10.insert(END,save_name_list[9])

    tkinter.Button(change_window,text="保存",font=mid_font,command=lambda:change_savelist(entry1.get(),entry2.get(),entry3.get(),entry4.get(),entry5.get(),entry6.get(),entry7.get(),entry8.get(),entry9.get(),entry10.get())).place(x=60,y=270)

def change_savelist(in1,in2,in3,in4,in5,in6,in7,in8,in9,in10):
    in_list=[in1,in2,in3,in4,in5,in6,in7,in8,in9,in10]
    try:
        load_preset5=load_workbook("preset.xlsx", data_only=True)
        db_custom2=load_preset5["custom"]
        
        for i in range(1,11): # 这里是1- 10 没有11 注意excel开始1 list开始1
            db_custom2.cell(i,5).value=in_list[i-1] # 这里保存了名字，直接导入
        global save_name_list
        save_name_list=in_list
        load_preset5.save("preset.xlsx")
        load_preset5.close()
        save_select.set(save_name_list[0])
        save_select['values']=save_name_list
        change_window.destroy() # 自己退出
        tkinter.messagebox.showinfo("通知","保存完成")
    except PermissionError as error:
        tkinter.messagebox.showerror("错误","请关闭重试")

# 这个作者一点都不加说明吗？？？
def update_count():
    global count_num, count_all, show_number # 这几个值是随时变动的全局变量
    global showcon,all_list_num # 这也是随时变动的全局变量
    while True:
        showcon(text=str(count_num)+"有效搭配/"+str(count_all)+"无效\n"+str(all_list_num)+"整体")
        time.sleep(0.1)

def update_count2(): # 这里都是计算数量
    while True:
        global select_item
        a_num_all=0 # 而且这个all都没用到
        a_num=[0,0,0,0,0,0,0,0,0,0,0] # 这个是用来存储啥的，看不太懂
        for i in range(101,136):
            try:
                a_num[0]=a_num[0]+select_item['tg1{}0'.format(i)]+select_item['tg1{}1'.format(i)] # 这里是因为普通上衣 和 神话上衣的重复计算
            except KeyError as error:
                p=0
            try:
                a_num[1]=a_num[1]+select_item['tg1{}0'.format(i+100)]
            except KeyError as error:
                p=0
            try:
                a_num[2]=a_num[2]+select_item['tg1{}0'.format(i+200)]
            except KeyError as error:
                p=0
            try:
                a_num[3]=a_num[3]+select_item['tg1{}0'.format(i+300)]
            except KeyError as error:
                p=0
            try:
                a_num[4]=a_num[4]+select_item['tg1{}0'.format(i+400)]
            except KeyError as error:
                p=0
            try:
                a_num[5]=a_num[5]+select_item['tg2{}0'.format(i)]+select_item['tg2{}1'.format(i)]
            except KeyError as error:
                p=0
            try:
                a_num[6]=a_num[6]+select_item['tg2{}0'.format(i+100)]
            except KeyError as error:
                p=0
            try:
                a_num[7]=a_num[7]+select_item['tg2{}0'.format(i+200)]
            except KeyError as error:
                p=0
            try:
                a_num[8]=a_num[8]+select_item['tg3{}0'.format(i)]
            except KeyError as error:
                p=0
            try:
                a_num[9]=a_num[9]+select_item['tg3{}0'.format(i+100)]
            except KeyError as error:
                p=0
            try:
                a_num[10]=a_num[10]+select_item['tg3{}0'.format(i+200)]+select_item['tg3{}1'.format(i+200)]
            except KeyError as error:
                p=0
                
        if a_num[0]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[1]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[2]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[3]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[4]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;

        if a_num[5]+a_num[6]+a_num[7]<2:
            a_num[5]=a_num[5]+1;a_num[6]=a_num[6]+1;a_num[7]=a_num[7]+1
        if a_num[8]+a_num[9]+a_num[10]<2:
            a_num[8]=a_num[8]+1;a_num[9]=a_num[9]+1;a_num[10]=a_num[10]+1
            
        if a_num[5]==0:
            a_num[5]=a_num[5]+1
        if a_num[6]==0:
            a_num[6]=a_num[6]+1
        if a_num[7]==0:
            a_num[7]=a_num[7]+1
            
        if a_num[8]==0:
            a_num[8]=a_num[8]+1
        if a_num[9]==0:
            a_num[9]=a_num[9]+1
        if a_num[10]==0:
            a_num[10]=a_num[10]+1
            
        a_num_all=a_num[0]*a_num[1]*a_num[2]*a_num[3]*a_num[4]*a_num[5]*a_num[6]*a_num[7]*a_num[8]*a_num[9]*a_num[10]
        showcon2(text="时数（1秒刷新）= "+str(a_num_all)) # 这里大体是排列组合算法算出多少种
        if a_num_all>10000000:
            show_count2['fg']="red" # 这里是直接修改属性
        else:
            show_count2['fg']="white" # 这里也是直接修改的属性
        time.sleep(1)

def update_thread():
    threading.Thread(target=update_count,daemon=True).start()
    threading.Thread(target=update_count2,daemon=True).start()

def timeline_select():
    global timeline_window
    timeline_window=tkinter.Toplevel(self)
    timeline_window.attributes("-topmost", True) 
    timeline_window.geometry("310x150+750+20")
    tkinter.Label(timeline_window,text="角色名=\n(准确的)",font=guide_font).place(x=10,y=9)
    cha_name=tkinter.Entry(timeline_window,width=13)
    cha_name.place(x=80,y=12)
    tkinter.Label(timeline_window,text="服务器名称",font=guide_font).place(x=10,y=59)
    sever_list=['카인','디레지에','巴卡尔','希勒','安东','卡西利亚斯','弗雷','希洛克']
    serv_name=tkinter.ttk.Combobox(timeline_window,values=sever_list,width=11)
    serv_name.place(x=80,y=62)
    serv_name.set('카인')
    load_timeline=tkinter.Button(timeline_window,command=lambda:show_timeline(cha_name.get(),serv_name.get()),text="불러오기",font=mid_font)
    load_timeline.place(x=200,y=25)
    tkinter.Label(timeline_window,text="在时间轴中仅加载史诗（第X部分）",fg="Red").place(x=10,y=100)
    tkinter.Label(timeline_window,text="如果服务器不稳定请多试几次",fg="Red").place(x=10,y=120)

def show_timeline(name,server):
    
    server_dict={'안톤':'anton','바칼':'bakal','카인':'cain','카시야스':'casillas',
                '디레지에':'diregie','힐더':'hilder','프레이':'prey','시로코':'siroco'}
    try:
        sever_code=server_dict[server]
        cha_id_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters?characterName='+parse.quote(name)+'&apikey=' + apikey)
        cha_id_dic=loads(cha_id_api.read().decode("utf-8"))
        cha_id=cha_id_dic['rows'][0]['characterId']

    ##
        print(sever_code)
        print(cha_id)
        time.sleep(0.3)
        start_time='20200101T0000'
        now=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        now2='20200101T0000'
        now3='20200101T0000'
        now4='20200101T0000'
        if int(now[0:6]) >= 202004:
            now='20200401T0000'
            now2=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        if int(now2[0:6]) >= 202007:
            now2='20200701T0000'
            now3=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        if int(now3[0:6]) >= 202010:
            now3='20201001T0000'
            now4=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        time_code='504,505,506,507,508,510,511,512,513,514'
        timeline_list=[]
        for nows in [[now,start_time],[now2,now],[now3,now2],[now4,now3]]:
            if nows[0] != '20200101T0000':
                timeline=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/timeline?limit=100&code='+time_code+'&startDate='+nows[1]+'&endDate='+nows[0]+'&apikey='+apikey)
                timeline2=loads(timeline.read().decode("utf-8"))['timeline']
                show_next=timeline2['next']
                timeline_list=timeline_list+timeline2['rows']
                time.sleep(0.3)
                while show_next != None:
                    timeline_next=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/timeline?next='+show_next+'&apikey='+apikey)
                    timeline_next2=loads(timeline_next.read().decode("utf-8"))['timeline']
                    timeline_list=timeline_list+timeline_next2['rows']
                    time.sleep(0.3)
                    show_next=timeline_next2['next']
            
        all_item=[]
        for now in timeline_list:
            item=now['data']['itemId']
            all_item.append(item)
        xl=openpyxl.load_workbook("DATA.xlsx", data_only=True)
        sh=xl['one']
        
        reset()
            
        for i in range(76,257):
            api_cod=sh.cell(i,40).value
            if all_item.count(api_cod) != 0:
                select_item['tg{}'.format(str(sh.cell(i,1).value))]=1
        xl.close()
        check_equipment()
        for i in range(101,136):
            check_set(i)
        timeline_window.destroy()
    except urllib.error.HTTPError as error:
        tkinter.messagebox.showerror("错误","API 访问失败（网络错误）")

def reset():
    know_list2=['13390150','22390240','23390450','33390750','21400340','31400540','32410650']
    for i in range(1101,3336): # 这里是所有的装备列表
        try:
            select_item['tg{}0'.format(i)]=0
        except KeyError as error:
            passss=1
        try:
            select_item['tg{}1'.format(i)]=0
        except KeyError as error:
            passss=1
    for i in know_list2:
        select_item['tg{}'.format(i)]=0
    check_equipment()
    for i in range(101,136):
        check_set(i)

def guide_speed():
    tkinter.messagebox.showinfo("准确度选择","快速=删除单一散件n中速=包括散件, 神话优先\n慢速=所有限制解除(非常慢)")

select_item={} # 选择的装备
def click_equipment(code): # 检查装备
    if eval("select_item['tg{}']".format(code))==0:
        eval('select_{}'.format(code))['image']=image_list[str(code)]
        select_item['tg'+str('{}'.format(code))]=1
    elif eval("select_item['tg{}']".format(code))==1:
        eval('select_{}'.format(code))['image']=image_list2[str(code)]
        select_item['tg'+str('{}'.format(code))]=0
    if len(str(code))==5:
        check_set(int('1'+str(code)[2:4]))

def check_equipment():
    know_list2=['13390150','22390240','23390450','33390750','21400340','31400540','32410650']
    for i in range(11010,33352):
        try:
            if eval("select_item['tg{}']".format(i))==0:
                eval('select_{}'.format(i))['image']=image_list2[str(i)]
            elif eval("select_item['tg{}']".format(i))==1:
                eval('select_{}'.format(i))['image']=image_list[str(i)]
        except KeyError as error:
            c=1
    for i in know_list2:
        try:
            if eval("select_item['tg{}']".format(i))==0:
                eval('select_{}'.format(i))['image']=image_list2[str(i)]
            elif eval("select_item['tg{}']".format(i))==1:
                eval('select_{}'.format(i))['image']=image_list[str(i)]
        except KeyError as error:
            c=1

def click_set(code): # code = 101
    code_add=code-100
    code_str=str(code)[1:3] #这里我需要做一个验证 这里是取两位后两位
    set_checked=0
    if code >=116: ##악세/특장/스까면
        if 116<= code <=119:
            for i in range(21,24): ## 악세부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 123>= code >=120:
            for i in range(31,34): ## 특장부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 131>= code >=128:
            for i in [11,22,31]: ## 상목보부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 127>= code >=124:
            for i in [12,21,32]: ## 하팔법부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 135>= code >=132:
            for i in [15,23,33]: ## 신반귀부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        if set_checked==3: ## 채택 숫자가 3이면
            for i in range(11,36): ##모든 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list2[str(i)+code_str+'0'] ##이미지도 오프로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=0 ##모든 체크를 0으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set2[str(code)] ##세트이미지도 오프로 바꿈
        else: ## 채택 숫자가 3미만이면
            for i in range(11,36): ##모든 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list[str(i)+code_str+'0'] ##이미지도 온으로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=1 ##모든 체크를 1으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set[str(code)] ##세트이미지도 온으로 바꿈

            
    else:
        for i in range(11,16): ## 방어구 부위에서
            try:
                if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                    set_checked=set_checked+1 ##그럼 변수에 +1을 더함
            except KeyError as error:
                c=1
        
        if set_checked==5: ## 채택 숫자가 5이면
            for i in range(11,16): ## 방어구 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list2[str(i)+code_str+'0'] ##이미지도 오프로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=0 ##모든 체크를 0으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set2[str(code)] ##세트이미지도 오프로 바꿈
            
        else: ## 채택 숫자가 5미만이면
            for i in range(11,16): ## 방어구 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list[str(i)+code_str+'0'] ##이미지도 온으로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=1 ##모든 체크를 1으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set[str(code)] ##세트이미지도 온으로 바꿈
                    
def check_set(code): # 检查设置
    code_str=str(code)[1:3]
    set_checked=0
    if code < 116:
        for i in [11,12,13,14,15]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 120:
        for i in [21,22,23]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 124:
        for i in [31,32,33]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 128:
        for i in [12,21,32]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 132:
        for i in [11,22,31]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 136:
        for i in [15,23,33]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
                
    if code < 116:
        set115
        if set_checked==5:
            eval('set'+str(code))['image']=image_list_set[str(code)] # 这里进行替换图片
        else:
            eval('set'+str(code))['image']=image_list_set2[str(code)]
    else:
        if set_checked==3:
            eval('set'+str(code))['image']=image_list_set[str(code)]
        else:
            eval('set'+str(code))['image']=image_list_set2[str(code)]

def cha_select(jobs):
    global cha_window
    cha_window=tkinter.Toplevel(self)
    cha_window.attributes("-topmost", True) 
    cha_window.geometry("350x200+750+20")
    tkinter.Label(cha_window,text="角色名称=\n(정확히)",font=guide_font).place(x=10,y=9)
    cha_name=tkinter.Entry(cha_window,width=15)
    cha_name.place(x=80,y=12)
    tkinter.Label(cha_window,text="서버명=",font=guide_font).place(x=10,y=49)
    sever_list=['카인','디레지에','바칼','힐더','안톤','卡西利亚斯','프레이','시로코']
    serv_name=tkinter.ttk.Combobox(cha_window,values=sever_list,width=13)
    serv_name.place(x=80,y=52)
    serv_name.set('카인')
    tkinter.Label(cha_window,text="职业=",font=guide_font).place(x=10,y=79)
    job_name=tkinter.ttk.Combobox(cha_window,values=jobs,width=13)
    job_name.place(x=80,y=82)
    job_name.set('职业选择')

    
    load_timeline=tkinter.Button(cha_window,command=lambda:calc_my_cha(cha_name.get(),serv_name.get(),job_name.get()),text="불러오기",font=mid_font)
    load_timeline.place(x=240,y=25)
    tkinter.Label(cha_window,text="한캐릭씩 볼것!",font=mid_font,fg="Red").place(x=205,y=76)
    tkinter.Label(cha_window,text="오직 12부위 장비만 불러옵니다.(마부/강화,칭호/클쳐,압타 X)",fg="Red").place(x=7,y=110)
    tkinter.Label(cha_window,text="메인창과 호환을 위해 커스텀은 전부 메인창을 따라갑니다.",fg="Red").place(x=7,y=130)
    tkinter.Label(cha_window,text="100제 에픽이 아니면 전부 100제 레전으로 처리합니다.",fg="Red").place(x=7,y=150)
    tkinter.Label(cha_window,text="如果服务器不稳定请多试几次",fg="Red").place(x=7,y=170)

def calc_my_cha(cha_name,serv_name,job_name):
    info_stat=0
    global show_cha_result
    try:
        show_cha_result.destroy()
    except NameError as error:
        pass
    server_dict={'안톤':'anton','바칼':'bakal','카인':'cain','카시야스':'casillas',
                '디레지에':'diregie','힐더':'hilder','프레이':'prey','시로코':'siroco'}
    try:
        sever_code=server_dict[serv_name]
        cha_id_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters?characterName='+parse.quote(cha_name)+'&apikey=' + apikey)
        cha_id_dic=loads(cha_id_api.read().decode("utf-8"))
        cha_id=cha_id_dic['rows'][0]['characterId']
        print(sever_code)
        print(cha_id)
        cha_image(cha_id,sever_code)
        time.sleep(0.3)

        cha_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/equip/equipment?apikey='+apikey)
        cha_api_dic=loads(cha_api.read().decode("utf-8"))
        cha_equ=cha_api_dic["equipment"]
        item_list=[]
        for i in range(0,20):
            try:
                if cha_equ[i]['slotName'] != '보조무기':
                    item_list.append(cha_equ[i]['itemName'])
                if cha_equ[i]['slotName'] == '무기':
                    wep_name=cha_equ[i]['itemName']
            except IndexError as error:
                pass

        ##싀칭박스
        if job_name[4:7] == '奶爸' or job_name[4:7] == '奶妈' or job_name[4:7] == '奶萝':
            switch_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/skill/buff/equip/equipment?apikey='+apikey)
            switch_api_dic=loads(switch_api.read().decode("utf-8"))
            swi_equ=switch_api_dic["skill"]['buff']['equipment']
            swi_list=[]
            for i in range(0,20):
                try:
                    if swi_equ[i]['slotName'] != '보조무기':
                        swi_list.append(swi_equ[i]['itemName'])
                    if swi_equ[i]['slotName'] == '무기':
                        swi_wep_name=swi_equ[i]['itemName']
                except IndexError as error:
                    pass
    except urllib.error.HTTPError as error:
        tkinter.messagebox.showerror("错误","API 访问失败（网络错误）")

    xl=openpyxl.load_workbook("DATA.xlsx", data_only=True)
    sh=xl['one']
    for_calc_list=[]
    for_calc_list2=[]
    for i in item_list:
        for j in range(1,257):
            if i==sh.cell(j,39).value:
                for_calc_list.append('{}'.format(str(sh.cell(j,1).value)))

    check_exist=[]
    wep_exist=0
    for i in for_calc_list:
        if len(i) != 6:
            check_exist.append(i[0:2])
        if len(i) == 6:
            wep_exist=1
    if wep_exist ==0:
        for_calc_list.append('111001')
        wep_name='흑천의 주인(검은 성전의 기억)'
        
    for i in [11,12,13,14,15]:
        if check_exist.count(str(i))==0:
            for_calc_list.append(str(i)+'360')
    for i in [21,22,23]:
        if check_exist.count(str(i))==0:
            for_calc_list.append(str(i)+'370')
    for i in [31,32,33]:
        if check_exist.count(str(i))==0:
            for_calc_list.append(str(i)+'380')
            
    calc_now=[]
    for i in for_calc_list:
        if len(i)==6:
            wep_num=[i]
        else:
            calc_now.append(i)
    ##스위칭박스
    if job_name[4:7] == '奶爸' or job_name[4:7] == '奶妈' or job_name[4:7] == '奶萝':
        check_exist2=[]
        wep_exist2=0
        for i in swi_list:
            for j in range(1,257):
                if i==sh.cell(j,39).value:
                    for_calc_list2.append('{}'.format(str(sh.cell(j,1).value)))
        for i in for_calc_list2:
            if len(i) != 6:
                check_exist2.append(i[0:2])
            if len(i) == 6:
                wep_exist2=1
                
        if wep_exist2 ==0:
            for_calc_list2.append('111001')
            wep_name2='흑천의 주인(검은 성전의 기억)'
            
        for i in [11,12,13,14,15]:
            if check_exist2.count(str(i))==0:
                for_calc_list2.append(str(i)+'360')
        for i in [21,22,23]:
            if check_exist2.count(str(i))==0:
                for_calc_list2.append(str(i)+'370')
        for i in [31,32,33]:
            if check_exist2.count(str(i))==0:
                for_calc_list2.append(str(i)+'380')
                
        calc_now2=[]
        for i in for_calc_list2:
            if len(i)==6:
                wep_num2=[i]
            else:
                calc_now2.append(i)

    c=1        
    db_buf=xl["buf"]
    opt_buf={}
    name_buf={}
    for row in db_buf.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_buf[db_buf.cell(c,1).value]=row_value_cut ## DB 불러오기 ##
        name_buf[db_buf.cell(c,1).value]=row_value
        c=c+1

    d=1        
    db_buflvl=xl["buflvl"]
    opt_buflvl={}
    for row in db_buflvl.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = [0] + row_value[1:]
        opt_buflvl[db_buflvl.cell(d,1).value]=row_value_cut
        d=d+1

    load_presetc=load_workbook("preset.xlsx", data_only=True)
    db_preset=load_presetc["custom"]
    ele_skill=int(opt_job_ele[job_name][1])
    ele_in=(int(db_preset["B14"].value)+int(db_preset["B15"].value)+int(db_preset["B16"].value)+
            int(ele_skill)-int(db_preset["B18"].value)+int(db_preset["B19"].value)+13)
    
    if job_name[-4:] == "(奶系)":
        active_eff_one=15
        active_eff_set=18-3
    else:
        active_eff_one=21
        active_eff_set=27-3
    global time_select
    if time_select.get() == "60秒(觉醒占比↓)":
        lvl_shift=6
    else:
        lvl_shift=0
    job_lv1=opt_job[job_name][11+lvl_shift]
    job_lv2=opt_job[job_name][12+lvl_shift]
    job_lv3=opt_job[job_name][13+lvl_shift]
    job_lv4=opt_job[job_name][14+lvl_shift]
    job_lv5=opt_job[job_name][15+lvl_shift]
    job_lv6=opt_job[job_name][16+lvl_shift]
    job_pas0=opt_job[job_name][0]
    job_pas1=opt_job[job_name][1]
    job_pas2=opt_job[job_name][2]
    job_pas3=opt_job[job_name][3]
    
    
    extra_dam=0
    extra_cri=0
    extra_bon=0
    extra_all=0
    extra_pas2=0
    extra_sta=0
    extra_att=0
    if style_select.get() == '伟大的意志':
        extra_cri=extra_cri+18
        extra_sta=extra_sta+4
    if style_select.get() == '使徒降临':
        extra_bon=extra_bon+12
        extra_sta=extra_sta+3
    if style_select.get() == '国庆称号':
        extra_cri=extra_cri+15 
    if creature_select.get() == '普通宠物':
        extra_bon=extra_bon+12
        extra_sta=extra_sta+10
    if creature_select.get() == '至尊宠物':
        extra_bon=extra_bon+15
        extra_sta=extra_sta+12
        
    
    set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)] 
    set_on=[];setapp=set_on.append
    setcount=set_list.count
    for i in range(101,136):
        if setcount(str(i))==2:
            setapp(str(i)+"1")
        if 4>=setcount(str(i))>=3:
            setapp(str(i)+"2")
        if setcount(str(i))==5:
            setapp(str(i)+"3")
    for i in range(136,139):
        if setcount(str(i))==2:
            setapp(str(i)+"0")
        if 4>=setcount(str(i))>=3:
            setapp(str(i)+"1")
        if setcount(str(i))==5:
            setapp(str(i)+"2")
    calc_wep=wep_num+calc_now
    for_calc=set_on+calc_wep

    if job_name[4:7] != '奶爸' and job_name[4:7] != '奶妈' and job_name[4:7] != '奶萝':
                       
        skiper=0
        damage=0
        base_array=np.array([0,0,extra_dam,extra_cri,extra_bon,0,extra_all,extra_att,extra_sta,ele_in,0,1,0,0,0,0,0,0,extra_pas2,0,0,0,0,0,0,0,0,0])
        oneone=len(for_calc)
        oneonelist=[]
        for i in range(oneone):
            no_cut=opt_one.get(for_calc[i])               ## 11번 스증
            cut=np.array(no_cut[0:20]+no_cut[22:23]+no_cut[34:35]+no_cut[38:44])
            skiper=(skiper/100+1)*(cut[11]/100+1)*100-100
            oneonelist.append(cut)
        for i in range(oneone):
            base_array=base_array+oneonelist[i]
        # 코드 이름
        # 0추스탯 1추공 2증 3크 4추 5속추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 二觉캐특수액티브 /22~27 액티브레벨링

        if set_on.count('1201')==1 and calc_wep.count('32200')==1:
            base_array[3]=base_array[3]-5
        if calc_wep.count('33230')==1 or calc_wep.count('33231')==1:
            if calc_wep.count('31230')==0:
                base_array[4]=base_array[4]-10
            if calc_wep.count('32230')==0:
                base_array[9]=base_array[9]-40
        if calc_wep.count('15340')==1 or calc_wep.count('23340')==1 or calc_wep.count('33340')==1 or calc_wep.count('33341')==1:
            if set_on.count('1341')==0 and set_on.count('1342') ==0:
                if calc_wep.count('15340')==1:
                    base_array[9]=base_array[9]-20
                elif calc_wep.count('23340')==1:
                    base_array[2]=base_array[2]-10
                elif calc_wep.count('33340')==1:
                    base_array[6]=base_array[6]-5
                else:
                    base_array[9]=base_array[9]-4
                    base_array[2]=base_array[2]-2
                    base_array[6]=base_array[6]-1
                    base_array[8]=base_array[8]-1.93
        if calc_wep.count('11111')==1:
            if set_on.count('1112')==1 or set_on.count('1113')==1:
                base_array[20]=base_array[20]+10
        if calc_wep.count('11301')==1:
            if calc_wep.count('22300')!=1:
                base_array[4]=base_array[4]-10
                base_array[7]=base_array[7]+10
            if calc_wep.count('31300')!=1:
                base_array[4]=base_array[4]-10
                base_array[7]=base_array[7]+10
                        
        real_bon=base_array[4]+base_array[5]*(base_array[9]*0.0045+1.05)
        actlvl=((base_array[active_eff_one]+base_array[22]*job_lv1+base_array[23]*job_lv2+base_array[24]*job_lv3+
                base_array[25]*job_lv4+base_array[26]*job_lv5+base_array[27]*job_lv6)/100+1)
        paslvl=((100+base_array[16]*job_pas0)/100)*((100+base_array[17]*job_pas1)/100)*((100+base_array[18]*job_pas2)/100)*((100+base_array[19]*job_pas3)/100)
        damage=((base_array[2]/100+1)*(base_array[3]/100+1)*(real_bon/100+1)*(base_array[6]/100+1)*(base_array[7]/100+1)*
                (base_array[8]/100+1)*(base_array[9]*0.0045+1.05)*(base_array[10]/100+1)*(skiper/100+1)*(base_array[12]/100+1)*
                actlvl*paslvl*((54500+3.31*base_array[0])/54500)*((4800+base_array[1])/4800)/(1.05+0.0045*int(ele_skill)))

        damage_with_cool=damage*(1+base_array[20]/100)
        type_code='deal'

        final_list=[damage,damage_with_cool]
        show_my_cha(calc_now,final_list,type_code,job_name,wep_name,ele_skill,cha_name,info_stat)
    else:

        set_list2=["1"+str(calc_now2[x][2:4]) for x in range(0,11)] 
        set_on2=[];setapp2=set_on2.append
        setcount2=set_list2.count
        for i in range(101,136):
            if setcount2(str(i))==2:
                setapp2(str(i)+"1")
            if 4>=setcount2(str(i))>=3:
                setapp2(str(i)+"2")
            if setcount2(str(i))==5:
                setapp2(str(i)+"3")
        for i in range(136,139):
            if setcount2(str(i))==2:
                setapp2(str(i)+"0")
            if 4>=setcount2(str(i))>=3:
                setapp2(str(i)+"1")
            if setcount2(str(i))==5:
                setapp2(str(i)+"2")
        calc_wep2=wep_num2+calc_now2
        for_calc2=set_on2+calc_wep2
        
        base_b=10+int(db_preset['H2'].value)+int(db_preset['H4'].value)+int(db_preset['H5'].value)+1
        base_c=12+int(db_preset['H3'].value)+1
        base_pas0=0
        base_pas0_c=3
        base_pas0_b=0
        base_stat_s=4114
        base_stat_h=4180
        base_pas0_1=0
        lvlget=opt_buflvl.get
        #코드 이름
        #0 체정 1 지능
        #祝福 2 스탯% 3 물공% 4 마공% 5 독공%
        #아포 6 고정 7 스탯%
        #8 축렙 9 포렙
        #10 아리아/보징증폭
        #11 전직패 12 보징 13 각패1 14 각패2 15 二觉 16 각패3
        #17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
        setget=opt_buf.get
        base_array=np.array([base_stat_h,base_stat_s,0,0,0,0,0,0,base_b,base_c,0,base_pas0,base_pas0_1,0,0,0,0,0,0,0,0])
        base_array2=np.array([base_stat_h,base_stat_s,0,0,0,0,0,0,base_b,base_c,0,base_pas0,base_pas0_1,0,0,0,0,0,0,0,0])
        calc_wep=wep_num+calc_now
        calc_wep2=wep_num2+calc_now2

        b_stat=10.24 ##탈리스만
        b_phy=0
        b_mag=0
        b_ind=0
        c_per=0
        b_stat2=10.24 ##탈리스만
        b_phy2=0
        b_mag2=0
        b_ind2=0
        c_per2=0
        oneone=len(for_calc)
        oneonelist=[]
        for i in range(oneone):
            no_cut=np.array(setget(for_calc[i]))             ## 2 3 4 5 7
            base_array=base_array+no_cut
            b_stat=(b_stat/100+1)*(no_cut[2]/100+1)*100-100
            b_phy=(b_phy/100+1)*(no_cut[3]/100+1)*100-100
            b_mag=(b_mag/100+1)*(no_cut[4]/100+1)*100-100
            b_ind=(b_ind/100+1)*(no_cut[5]/100+1)*100-100
            c_per=(c_per/100+1)*(no_cut[7]/100+1)*100-100
            oneonelist.append(no_cut)
            
        oneone2=len(for_calc2)
        oneonelist2=[]
        for i in range(oneone2):
            no_cut2=np.array(setget(for_calc2[i]))             ## 2 3 4 5 7
            base_array2=base_array2+no_cut2
            b_stat2=(b_stat2/100+1)*(no_cut2[2]/100+1)*100-100
            b_phy2=(b_phy2/100+1)*(no_cut2[3]/100+1)*100-100
            b_mag2=(b_mag2/100+1)*(no_cut2[4]/100+1)*100-100
            b_ind2=(b_ind2/100+1)*(no_cut2[5]/100+1)*100-100
            c_per2=(c_per2/100+1)*(no_cut2[7]/100+1)*100-100
            oneonelist2.append(no_cut)
                        
        #0 체정 1 지능
        #祝福 2 스탯% 3 물공% 4 마공% 5 독공%
        #아포 6 고정 7 스탯%
        #8 축렙 9 포렙
        #10 아리아/보징증폭
        #11 전직패 12 보징 13 각패1 14 각패2 15 二觉 16 각패3
        #17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
        if job_name[4:7] == "奶爸":
            b_base_att=lvlget('hol_b_atta')[int(base_array2[8])]#
            stat_pas0lvl_b=lvlget('pas0')[int(base_array2[11])+base_pas0_b]+lvlget('hol_pas0_1')[int(base_array2[12])]#
            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+base_pas0_c]+lvlget('hol_pas0_1')[int(base_array[12])]
            stat_pas1lvl=lvlget('hol_pas1')[int(base_array[13])]
            stat_pas1lvl2=lvlget('hol_pas1')[int(base_array2[13])]#
            stat_pas2lvl=lvlget('hol_act2')[int(base_array[15])]
            stat_pas2lvl2=lvlget('hol_act2')[int(base_array2[15])]#
            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
            stat_pas3lvl2=lvlget('pas3')[int(base_array2[16])]#
            stat_b=base_array2[0]+stat_pas0lvl_b+stat_pas1lvl2+stat_pas2lvl2+stat_pas3lvl2+19*base_array2[10]+int(db_preset['H6'].value)
            stat_c=base_array[0]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]+int(db_preset['H1'].value)
            b_stat_calc=int(int(lvlget('hol_b_stat')[int(base_array2[8])]*(b_stat2/100+1))*(stat_b/630+1))
            b_phy_calc=int(int(b_base_att*(b_phy2/100+1))*(stat_b/630+1))
            b_mag_calc=int(int(b_base_att*(b_mag2/100+1))*(stat_b/630+1))
            b_ind_calc=int(int(b_base_att*(b_ind2/100+1))*(stat_b/630+1))
            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
            pas1_calc=int(lvlget('hol_pas1_out')[int(base_array[13])]+213+base_array[17])
            pas1_out=str(int(lvlget('hol_pas1_out')[int(base_array[13])]+213+base_array[17]))+"  ("+str(int(20+base_array[13]))+"렙)"
            save1=str(b_stat_calc)+"/"+str(b_average)+"   ["+str(int(stat_b))+"("+str(int(base_array2[8]))+"렙)]"
            info_stat=int(stat_c-stat_pas2lvl-528-stat_pas1lvl-213-lvlget('hol_pas0_1')[int(base_array[12])])

        else:
            if job_name[4:7] == "奶妈":
                b_value=675
                aria=1.25+0.05*base_array2[10]
            if job_name[4:7] == "奶萝":
                b_value=665
                aria=(1.20+0.05*base_array2[10])*1.20
                            
            b_base_att=lvlget('se_b_atta')[int(base_array2[8])]#
            stat_pas0lvl_b=lvlget('pas0')[int(base_array2[11])+int(base_pas0_b)]#
            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+int(base_pas0_c)]
            stat_pas1lvl=lvlget('se_pas1')[int(base_array[13])]+base_array[18]
            stat_pas1lvl2=lvlget('se_pas1')[int(base_array2[13])]+base_array2[18]#
            stat_pas2lvl=lvlget('se_pas2')[int(base_array[14])]
            stat_pas2lvl2=lvlget('se_pas2')[int(base_array2[14])]#
            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
            stat_pas3lvl2=lvlget('pas3')[int(base_array2[16])]#
            stat_b=base_array2[1]+stat_pas0lvl_b+stat_pas1lvl2+stat_pas2lvl2+stat_pas3lvl2+int(db_preset['H6'].value)
            stat_c=base_array[1]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+int(db_preset['H1'].value)
            b_stat_calc=int(int(lvlget('se_b_stat')[int(base_array2[8])]*(b_stat2/100+1))*(stat_b/b_value+1)*aria)
            b_phy_calc=int(int(b_base_att*(b_phy2/100+1)*(stat_b/b_value+1))*aria)
            b_mag_calc=int(int(b_base_att*(b_mag2/100+1)*(stat_b/b_value+1))*aria)
            b_ind_calc=int(int(b_base_att*(b_ind2/100+1)*(stat_b/b_value+1))*aria)
            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
            pas1_calc=int(stat_pas1lvl+442)
            pas1_out=str(int(stat_pas1lvl+442))+"  ("+str(int(20+base_array[13]))+"렙)"
            save1=str(b_stat_calc)+"("+str(int(b_stat_calc/aria))+")/"+str(b_average)+"("+str(int(b_average/aria))+")\n                  ["+str(int(stat_b))+"("+str(int(base_array2[8]))+"렙)]"
            info_stat=stat_c-pas1_calc
        ##1축 2포 3합
        save2=str(c_calc)+"    ["+str(int(stat_c))+"("+str(int(base_array[9]))+"렙)]"
        save3=pas1_out
        save4=((15000+pas1_calc+c_calc+b_stat_calc)/250+1)*(2650+b_average)
        type_code='buf'
        final_list=[save1,save2,save3,save4]
        show_my_cha(calc_now,final_list,type_code,job_name,wep_name,ele_skill,cha_name,info_stat,calc_now2)

    xl.close()
    load_presetc.close()
    
def cha_image(cha_code,serv_code):
    down_url='https://img-api.neople.co.kr/df/servers/'+serv_code+'/characters/'+cha_code+'?zoom=1?apikey='+apikey
    urllib.request.urlretrieve(down_url,'my_cha.png')

def show_my_cha(equipment,final_list,type_code,job_name,wep_name,ele_skill,cha_name,info_stat,*equipment2):
    global show_cha_result
    show_cha_result=Toplevel(self)
    show_cha_result.geometry("244x402+750+320")
    show_cha_result.resizable(False,False)
    global canvas,cha_bg,cha_img
    canvas = Canvas(show_cha_result, width=246, height=404, bd=0)
    canvas.place(x=-2,y=-2)
    cha_bg=tkinter.PhotoImage(file='ext_img/bg_cha_img.png')
    canvas.create_image(123,202,image=cha_bg)
    cha_img=tkinter.PhotoImage(file='my_cha.png')
    canvas.create_image(123,70,image=cha_img)
    global image_list,image_on
    image_on={}
    for i in [11,12,13,14,15,21,22,23,31,32,33]:
        for j in equipment:
            if j[0:2] == str(i):
                image_on[str(i)]=image_list[j]
    global img11,img12,img13,img14,img15,img21,img22,img23,img31,img32,img33
    img11=canvas.create_image(57,57,image=image_on['11'])
    img12=canvas.create_image(27,87,image=image_on['12'])
    img13=canvas.create_image(27,57,image=image_on['13'])
    img14=canvas.create_image(57,87,image=image_on['14'])
    img15=canvas.create_image(27,117,image=image_on['15'])
    img21=canvas.create_image(189,57,image=image_on['21'])
    img22=canvas.create_image(219,57,image=image_on['22'])
    img23=canvas.create_image(219,87,image=image_on['23'])
    img31=canvas.create_image(189,87,image=image_on['31'])
    img32=canvas.create_image(219,117,image=image_on['32'])
    img33=canvas.create_image(189,117,image=image_on['33'])

    canvas.create_text(123,197,text="职业= "+job_name,font=guide_font,fill='white')
    canvas.create_text(123,217,text="武器= "+wep_name,font=guide_font,fill='white')
    canvas.create_text(58,157,text=cha_name,font=guide_font,fill='white')
    


    if type_code =='deal':
        canvas.create_text(123,237,text="技能快速（反向纠正率）= "+str(ele_skill)+' (-'+str(int(100*(1-1.05/(1.05+int(ele_skill)*0.0045))))+"%)",font=guide_font,fill='white')
        global style_select, creature_select, time_select
        canvas.create_text(123,257,text="延迟设置= "+time_select.get(),font=guide_font,fill='white')
        canvas.create_text(123,277,text=style_select.get()+" / "+creature_select.get(),font=guide_font,fill='white')
        canvas.create_text(123,307,text="纯伤害= "+str(int(100*final_list[0]))+"%",font=mid_font,fill='white')
        canvas.create_text(123,342,text="冷却补正= "+str(int(100*final_list[1]))+"%",font=mid_font,fill='white')

        canvas.create_text(123,377,text="请在主窗口中选择标题/生物.",font=guide_font,fill='red')

    if type_code =='buf':
        print(equipment2)
        global image_on2
        image_on2={}
        for i in [11,12,13,14,15,21,22,23,31,32,33]:
            for j in equipment2[0]:
                if j[0:2] == str(i):
                    image_on2[str(i)]=image_list[j]
        
        canvas.create_text(123,247,text="祝福="+final_list[0],font=guide_font,fill='white')
        canvas.create_text(123-17,282,text="一觉="+final_list[1],font=guide_font,fill='white')
        canvas.create_text(123-44,312,text="一觉패="+final_list[2],font=guide_font,fill='white')
        canvas.create_text(123,340,text="总增益技能= "+str(int(final_list[3]/10)),font=mid_font,fill='white')
        canvas.create_text(123,368,text="估计统计="+str(int(info_stat)),font=guide_font,fill='white')
        canvas.create_text(123,387,text="自定义统计/技能表",font=guide_font,fill='red')

        global toggle_cha, toggle_cha_but,toggle_swi_img,toggle_swi_img2
        toggle_cha=0
        toggle_swi_img=tkinter.PhotoImage(file='ext_img/info_swi.png')
        toggle_swi_img2=tkinter.PhotoImage(file='ext_img/info_swi2.png')
        toggle_cha_but=tkinter.Button(show_cha_result,borderwidth=0,activebackground=dark_main,bg=dark_main,image=toggle_swi_img,command=toggle_cha_swi)
        toggle_cha_but.image=toggle_swi_img
    
    canvas.image = cha_bg,cha_img,image_on['11'],image_on['12'],image_on['13'],image_on['14'],image_on['15'],image_on['21'],image_on['22'],image_on['23'],image_on['31'],image_on['32'],image_on['33']
    
    if type_code =='buf':
        toggle_cha_but.place(x=155,y=142)
        tkinter.messagebox.showinfo("小心！，缓冲区必须与“集成定制”中的统计信息和技能水平相匹配。 目的是要确定一个适合您的参考值，以计算缓冲区\ N \ n请注意，在测量唯一流的增强力时，将使用求解器部分！)")


def toggle_cha_swi():
    global toggle_cha,cha_bg,cha_img, toggle_cha_but,toggle_swi_img,toggle_swi_img2
    global image_on,image_on2
    if toggle_cha == 0:
        image_ont=image_on2
        toggle_cha_but['image']=toggle_swi_img2
        toggle_cha=1
    elif toggle_cha == 1:
        image_ont=image_on
        toggle_cha_but['image']=toggle_swi_img
        toggle_cha=0
    global img11,img12,img13,img14,img15,img21,img22,img23,img31,img32,img33
    canvas.itemconfig(img11,image=image_ont['11'])
    canvas.itemconfig(img12,image=image_ont['12'])
    canvas.itemconfig(img13,image=image_ont['13'])
    canvas.itemconfig(img14,image=image_ont['14'])
    canvas.itemconfig(img15,image=image_ont['15'])
    canvas.itemconfig(img21,image=image_ont['21'])
    canvas.itemconfig(img22,image=image_ont['22'])
    canvas.itemconfig(img23,image=image_ont['23'])
    canvas.itemconfig(img31,image=image_ont['31'])
    canvas.itemconfig(img32,image=image_ont['32'])
    canvas.itemconfig(img33,image=image_ont['33'])
    canvas.image = cha_bg,cha_img,image_ont['11'],image_ont['12'],image_ont['13'],image_ont['14'],image_ont['15'],image_ont['21'],image_ont['22'],image_ont['23'],image_ont['31'],image_ont['32'],image_ont['33']


def stop_calc():
    global exit_calc # 这个是全局标志位
    exit_calc=1
    time.sleep(1)
    exit_calc=0


    
## 内部结构 ## # 这是韩械写的内部结构
know_list=['13390150','22390240','23390450','33390750','21400340','31400540','32410650'] # 这几个是只会常务
image_list={} # 1 是明的
image_list2={} # 2 是暗的
image_list_set={} # 存储名字  只放名字
image_list_set2={} # 暗名字图片
for i in know_list:
    image_list[i]=eval('PhotoImage(file="image/{}n.png")'.format(i))
    image_list2[i]=eval('PhotoImage(file="image/{}f.png")'.format(i))
for i in range(1101,3339): # 从这里都给导进去
    try:
        image_list[str(i)+"0"]=eval('PhotoImage(file="image/{}0n.png")'.format(i))
    except TclError as error:
        passss=1
    try:
        image_list2[str(i)+"0"]=eval('PhotoImage(file="image/{}0f.png")'.format(i))
    except TclError as error:
        passss=1
    try:
        image_list[str(i)+"1"]=eval('PhotoImage(file="image/{}1n.gif")'.format(i))
        image_list2[str(i)+"1"]=eval('PhotoImage(file="image/{}1f.png")'.format(i))
    except TclError as error:
        passss=1
for i in range(1,36): # 这里是导入名字
    image_list_set[str(100+i)]=eval('PhotoImage(file="set_name/{}.png")'.format(i+100))
    image_list_set2[str(100+i)]=eval('PhotoImage(file="set_name/{}f.png")'.format(i+100))



    
select_perfect=tkinter.ttk.Combobox(self,values=['快速','中速','慢速'],width=15)
select_perfect.place(x=145,y=11)
select_perfect.set('中速')
select_speed_img=PhotoImage(file="ext_img/select_speed.png") # 弹框提示
tkinter.Button(self,command=guide_speed,image=select_speed_img,borderwidth=0,activebackground=dark_main,bg=dark_main).place(x=29,y=7)
reset_img=PhotoImage(file="ext_img/reset.png") # 这里是初始化 那块
tkinter.Button(self,command=reset,image=reset_img,borderwidth=0,activebackground=dark_main,bg=dark_main).place(x=302,y=476)


# 这里是填入武器的下拉菜单
wep_list=[] # 这个只存储名字
for i in range(0,75):
    wep_list.append(name_one[str(i+111001)][1]) # name_one 是一个字典
    
wep_image=PhotoImage(file="ext_img/wep.png")
wep_g=tkinter.Label(self,image=wep_image,borderwidth=0,activebackground=dark_main,bg=dark_main)
wep_g.place(x=29,y=55)
wep_select=tkinter.ttk.Combobox(self,width=30,values=wep_list) # 选择武器的下拉菜单
wep_select.place(x=110,y=60)
wep_select.set('选择武器') # 这里是默认值

time_select=tkinter.ttk.Combobox(self,width=13,values=['20秒(觉醒占比↑)','60秒(觉醒占比↓)']) # 下拉菜单，这里都是用的列表
time_select.set('20秒(觉醒占比↑)') # 下拉菜单，这里是设置默认值
time_select.place(x=390-17,y=220+52)
jobup_select=tkinter.ttk.Combobox(self,width=13,values=jobs)
jobup_select.set('职业选择') # 下拉菜单
jobup_select.place(x=390-17,y=190+52)
style_list=['伟大的意志','使徒降临','国庆称号','其他（直接比较）'] # 下拉菜单
style_select=tkinter.ttk.Combobox(self,width=13,values=style_list)
style_select.set('使徒降临') # 下拉菜单
style_select.place(x=390-17,y=250+52)
creature_list=['至尊宠物','普通宠物','其他（直接比较）'] # 下拉菜单
creature_select=tkinter.ttk.Combobox(self,width=13,values=creature_list)
creature_select.set('普通宠物') # 下拉菜单
creature_select.place(x=390-17,y=280+52)
req_cool=tkinter.ttk.Combobox(self,width=13,values=['X(纯伤害)','O(打开)'])
req_cool.set('X(纯伤害)') # 下来菜单
req_cool.place(x=390-17,y=310+52)

calc_img=PhotoImage(file="ext_img/calc.png") # 计算
select_all=tkinter.Button(self,image=calc_img,borderwidth=0,activebackground=dark_main,command=calc_thread,bg=dark_main)
select_all.place(x=390-35,y=7)
stop_img=PhotoImage(file="ext_img/stop.png") # 停止计算
tkinter.Button(self,image=stop_img,borderwidth=0,activebackground=dark_main,command=stop_calc,bg=dark_main).place(x=390-35,y=62)

timeline_img=PhotoImage(file="ext_img/timeline.png") # 时间线
select_custom=tkinter.Button(self,image=timeline_img,borderwidth=0,activebackground=dark_main,command=timeline_select,bg=dark_sub)
select_custom.place(x=345+165,y=340-100)
custom_img=PhotoImage(file="ext_img/custom.png") # 统一自定义
select_custom2=tkinter.Button(self,image=custom_img,borderwidth=0,activebackground=dark_main,command=costum,bg=dark_sub)
select_custom2.place(x=435+165,y=340-100)

save_select=tkinter.ttk.Combobox(self,width=8,values=save_name_list) # 这里必须加上ttk
save_select.place(x=345+165,y=410-100);save_select.set(save_name_list[0])
save_img=PhotoImage(file="ext_img/SAVE.png") #存档
save=tkinter.Button(self,image=save_img,borderwidth=0,activebackground=dark_main,command=save_checklist,bg=dark_sub)
save.place(x=345+165,y=440-100)
load_img=PhotoImage(file="ext_img/LOAD.png") # 读档
load=tkinter.Button(self,image=load_img,borderwidth=0,activebackground=dark_main,command=load_checklist,bg=dark_sub)
load.place(x=435+165,y=440-100)
change_name_img=PhotoImage(file="ext_img/name_change.png")
change_list_but=tkinter.Button(self,image=change_name_img,borderwidth=0,activebackground=dark_main,command=change_list_name,bg=dark_sub)
change_list_but.place(x=435+165,y=405-100)
calc_me_img=PhotoImage(file="ext_img/calc_me.png") # 查看自我角色那个图片
calc_me=tkinter.Button(self,image=calc_me_img,borderwidth=0,activebackground=dark_main,command=lambda:cha_select(jobs),bg=dark_sub)
calc_me.place(x=300,y=400)


show_count=tkinter.Label(self,font=guide_font,fg="white",bg=dark_sub)
show_count.place(x=490,y=40)
showcon=show_count.configure
show_state=tkinter.Label(self,text="计算栏",font=guide_font,fg="white",bg=dark_sub)
show_state.place(x=490,y=20)
showsta=show_state.configure

show_count2=tkinter.Label(self,font=guide_font,fg="white",bg=dark_sub)
show_count2.place(x=430,y=480)
showcon2=show_count2.configure  # 这里showcon2 就代表 configure 函数

# 放装备的名字
set101=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['101'],command=lambda:click_set(101));set101.place(x=29,y=100)
set102=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['102'],command=lambda:click_set(102));set102.place(x=29,y=130)
set103=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['103'],command=lambda:click_set(103));set103.place(x=29,y=160)
set104=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['104'],command=lambda:click_set(104));set104.place(x=29,y=190)
set105=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['105'],command=lambda:click_set(105));set105.place(x=29,y=220)
set106=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['106'],command=lambda:click_set(106));set106.place(x=29,y=250)
set107=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['107'],command=lambda:click_set(107));set107.place(x=29,y=280)
set108=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['108'],command=lambda:click_set(108));set108.place(x=29,y=310)
set109=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['109'],command=lambda:click_set(109));set109.place(x=29,y=340)
set110=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['110'],command=lambda:click_set(110));set110.place(x=29,y=370)
set111=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['111'],command=lambda:click_set(111));set111.place(x=29,y=400)
set112=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['112'],command=lambda:click_set(112));set112.place(x=29,y=430)
set113=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['113'],command=lambda:click_set(113));set113.place(x=29,y=460)
set114=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['114'],command=lambda:click_set(114));set114.place(x=29,y=490)
set115=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['115'],command=lambda:click_set(115));set115.place(x=29,y=520) ##
set116=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['116'],command=lambda:click_set(116));set116.place(x=320-33,y=100)
set117=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['117'],command=lambda:click_set(117));set117.place(x=320-33,y=130)
set118=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['118'],command=lambda:click_set(118));set118.place(x=320-33,y=160)
set119=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['119'],command=lambda:click_set(119));set119.place(x=320-33,y=190) ##
set120=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['120'],command=lambda:click_set(120));set120.place(x=500-17,y=100)
set121=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['121'],command=lambda:click_set(121));set121.place(x=500-17,y=130)
set122=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['122'],command=lambda:click_set(122));set122.place(x=500-17,y=160)
set123=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['123'],command=lambda:click_set(123));set123.place(x=500-17,y=190) ##
set128=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['128'],command=lambda:click_set(128));set128.place(x=29,y=570)
set129=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['129'],command=lambda:click_set(129));set129.place(x=29,y=600)
set130=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['130'],command=lambda:click_set(130));set130.place(x=29,y=630)
set131=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['131'],command=lambda:click_set(131));set131.place(x=29,y=660) ##
set124=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['124'],command=lambda:click_set(124));set124.place(x=225,y=570)
set125=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['125'],command=lambda:click_set(125));set125.place(x=225,y=600)
set126=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['126'],command=lambda:click_set(126));set126.place(x=225,y=630)
set127=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['127'],command=lambda:click_set(127));set127.place(x=225,y=660) ##
set132=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['132'],command=lambda:click_set(132));set132.place(x=421,y=570)
set133=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['133'],command=lambda:click_set(133));set133.place(x=421,y=600)
set134=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['134'],command=lambda:click_set(134));set134.place(x=421,y=630)
set135=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['135'],command=lambda:click_set(135));set135.place(x=421,y=660) ##



##지혜의 산물 智慧的产物  这里放的都是装备了
know_image=PhotoImage(file="set_name/know_name.png") # 智慧的产物
tkinter.Label(self,bg=dark_main,image=know_image).place(x=302,y=520)
select_item['tg13390150']=0;select_13390150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13390150'],command=lambda:click_equipment(13390150))
select_13390150.place(x=403,y=520)
select_item['tg22390240']=0;select_22390240=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22390240'],command=lambda:click_equipment(22390240))
select_22390240.place(x=433,y=520)
select_item['tg23390450']=0;select_23390450=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23390450'],command=lambda:click_equipment(23390450))
select_23390450.place(x=463,y=520)
select_item['tg33390750']=0;select_33390750=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33390750'],command=lambda:click_equipment(33390750))
select_33390750.place(x=493,y=520)

select_item['tg21400340']=0;select_21400340=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21400340'],command=lambda:click_equipment(21400340))
select_21400340.place(x=524,y=520)
select_item['tg31400540']=0;select_31400540=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31400540'],command=lambda:click_equipment(31400540))
select_31400540.place(x=554,y=520)
select_item['tg32410650']=0;select_32410650=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32410650'],command=lambda:click_equipment(32410650))
select_32410650.place(x=584,y=520)

## 只包括前2件装备 上衣 神话
select_item['tg11010']=0;select_11010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11010'],command=lambda:click_equipment(11010))
select_11010.place(x=100,y=100)
select_item['tg11011']=0;select_11011=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11011'],command=lambda:click_equipment(11011))
select_11011.place(x=130,y=100)
select_item['tg11020']=0;select_11020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11020'],command=lambda:click_equipment(11020))
select_11020.place(x=100,y=130)
select_item['tg11021']=0;select_11021=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11021'],command=lambda:click_equipment(11021))
select_11021.place(x=130,y=130)
select_item['tg11030']=0;select_11030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11030'],command=lambda:click_equipment(11030))
select_11030.place(x=100,y=160)
select_item['tg11031']=0;select_11031=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11031'],command=lambda:click_equipment(11031))
select_11031.place(x=130,y=160)
select_item['tg11040']=0;select_11040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11040'],command=lambda:click_equipment(11040))
select_11040.place(x=100,y=190)
select_item['tg11041']=0;select_11041=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11041'],command=lambda:click_equipment(11041))
select_11041.place(x=130,y=190)
select_item['tg11050']=0;select_11050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11050'],command=lambda:click_equipment(11050))
select_11050.place(x=100,y=220)
select_item['tg11051']=0;select_11051=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11051'],command=lambda:click_equipment(11051))
select_11051.place(x=130,y=220)
select_item['tg11060']=0;select_11060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11060'],command=lambda:click_equipment(11060))
select_11060.place(x=100,y=250)
select_item['tg11061']=0;select_11061=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11061'],command=lambda:click_equipment(11061))
select_11061.place(x=130,y=250)
select_item['tg11070']=0;select_11070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11070'],command=lambda:click_equipment(11070))
select_11070.place(x=100,y=280)
select_item['tg11071']=0;select_11071=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11071'],command=lambda:click_equipment(11071))
select_11071.place(x=130,y=280)
select_item['tg11080']=0;select_11080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11080'],command=lambda:click_equipment(11080))
select_11080.place(x=100,y=310)
select_item['tg11081']=0;select_11081=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11081'],command=lambda:click_equipment(11081))
select_11081.place(x=130,y=310)
select_item['tg11090']=0;select_11090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11090'],command=lambda:click_equipment(11090))
select_11090.place(x=100,y=340)
select_item['tg11091']=0;select_11091=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11091'],command=lambda:click_equipment(11091))
select_11091.place(x=130,y=340)
select_item['tg11100']=0;select_11100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11100'],command=lambda:click_equipment(11100))
select_11100.place(x=100,y=370)
select_item['tg11101']=0;select_11101=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11101'],command=lambda:click_equipment(11101))
select_11101.place(x=130,y=370)
select_item['tg11110']=0;select_11110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11110'],command=lambda:click_equipment(11110))
select_11110.place(x=100,y=400)
select_item['tg11111']=0;select_11111=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11111'],command=lambda:click_equipment(11111))
select_11111.place(x=130,y=400)
select_item['tg11120']=0;select_11120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11120'],command=lambda:click_equipment(11120))
select_11120.place(x=100,y=430)
select_item['tg11121']=0;select_11121=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11121'],command=lambda:click_equipment(11121))
select_11121.place(x=130,y=430)
select_item['tg11130']=0;select_11130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11130'],command=lambda:click_equipment(11130))
select_11130.place(x=100,y=460)
select_item['tg11131']=0;select_11131=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11131'],command=lambda:click_equipment(11131))
select_11131.place(x=130,y=460)
select_item['tg11140']=0;select_11140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11140'],command=lambda:click_equipment(11140))
select_11140.place(x=100,y=490)
select_item['tg11141']=0;select_11141=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11141'],command=lambda:click_equipment(11141))
select_11141.place(x=130,y=490)
select_item['tg11150']=0;select_11150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11150'],command=lambda:click_equipment(11150))
select_11150.place(x=100,y=520)
select_item['tg11151']=0;select_11151=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11151'],command=lambda:click_equipment(11151))
select_11151.place(x=130,y=520)

# 这里到散搭
select_item['tg11280']=0;select_11280=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11280'],command=lambda:click_equipment(11280))
select_11280.place(x=100,y=570)
select_item['tg11281']=0;select_11281=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11281'],command=lambda:click_equipment(11281))
select_11281.place(x=130,y=570)
select_item['tg11290']=0;select_11290=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11290'],command=lambda:click_equipment(11290))
select_11290.place(x=100,y=600)
select_item['tg11291']=0;select_11291=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11291'],command=lambda:click_equipment(11291))
select_11291.place(x=130,y=600)
select_item['tg11300']=0;select_11300=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11300'],command=lambda:click_equipment(11300))
select_11300.place(x=100,y=630)
select_item['tg11301']=0;select_11301=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11301'],command=lambda:click_equipment(11301))
select_11301.place(x=130,y=630)
select_item['tg11310']=0;select_11310=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11310'],command=lambda:click_equipment(11310))
select_11310.place(x=100,y=660)
select_item['tg11311']=0;select_11311=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11311'],command=lambda:click_equipment(11311))
select_11311.place(x=130,y=660)
## 下装
select_item['tg12010']=0;select_12010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12010'],command=lambda:click_equipment(12010))
select_12010.place(x=161,y=100)
select_item['tg12020']=0;select_12020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12020'],command=lambda:click_equipment(12020))
select_12020.place(x=161,y=130)
select_item['tg12030']=0;select_12030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12030'],command=lambda:click_equipment(12030))
select_12030.place(x=161,y=160)
select_item['tg12040']=0;select_12040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12040'],command=lambda:click_equipment(12040))
select_12040.place(x=161,y=190)
select_item['tg12050']=0;select_12050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12050'],command=lambda:click_equipment(12050))
select_12050.place(x=161,y=220)
select_item['tg12060']=0;select_12060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12060'],command=lambda:click_equipment(12060))
select_12060.place(x=161,y=250)
select_item['tg12070']=0;select_12070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12070'],command=lambda:click_equipment(12070))
select_12070.place(x=161,y=280)
select_item['tg12080']=0;select_12080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12080'],command=lambda:click_equipment(12080))
select_12080.place(x=161,y=310)
select_item['tg12090']=0;select_12090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12090'],command=lambda:click_equipment(12090))
select_12090.place(x=161,y=340)
select_item['tg12100']=0;select_12100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12100'],command=lambda:click_equipment(12100))
select_12100.place(x=161,y=370)
select_item['tg12110']=0;select_12110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12110'],command=lambda:click_equipment(12110))
select_12110.place(x=161,y=400)
select_item['tg12120']=0;select_12120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12120'],command=lambda:click_equipment(12120))
select_12120.place(x=161,y=430)
select_item['tg12130']=0;select_12130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12130'],command=lambda:click_equipment(12130))
select_12130.place(x=161,y=460)
select_item['tg12140']=0;select_12140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12140'],command=lambda:click_equipment(12140))
select_12140.place(x=161,y=490)
select_item['tg12150']=0;select_12150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12150'],command=lambda:click_equipment(12150))
select_12150.place(x=161,y=520)
select_item['tg12240']=0;select_12240=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12240'],command=lambda:click_equipment(12240))
select_12240.place(x=296,y=570)
select_item['tg12250']=0;select_12250=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12250'],command=lambda:click_equipment(12250))
select_12250.place(x=296,y=600)
select_item['tg12260']=0;select_12260=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12260'],command=lambda:click_equipment(12260))
select_12260.place(x=296,y=630)
select_item['tg12270']=0;select_12270=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12270'],command=lambda:click_equipment(12270))
select_12270.place(x=296,y=660)
## 头肩
select_item['tg13010']=0;select_13010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13010'],command=lambda:click_equipment(13010))
select_13010.place(x=192,y=100)
select_item['tg13020']=0;select_13020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13020'],command=lambda:click_equipment(13020))
select_13020.place(x=192,y=130)
select_item['tg13030']=0;select_13030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13030'],command=lambda:click_equipment(13030))
select_13030.place(x=192,y=160)
select_item['tg13040']=0;select_13040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13040'],command=lambda:click_equipment(13040))
select_13040.place(x=192,y=190)
select_item['tg13050']=0;select_13050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13050'],command=lambda:click_equipment(13050))
select_13050.place(x=192,y=220)
select_item['tg13060']=0;select_13060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13060'],command=lambda:click_equipment(13060))
select_13060.place(x=192,y=250)
select_item['tg13070']=0;select_13070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13070'],command=lambda:click_equipment(13070))
select_13070.place(x=192,y=280)
select_item['tg13080']=0;select_13080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13080'],command=lambda:click_equipment(13080))
select_13080.place(x=192,y=310)
select_item['tg13090']=0;select_13090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13090'],command=lambda:click_equipment(13090))
select_13090.place(x=192,y=340)
select_item['tg13100']=0;select_13100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13100'],command=lambda:click_equipment(13100))
select_13100.place(x=192,y=370)
select_item['tg13110']=0;select_13110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13110'],command=lambda:click_equipment(13110))
select_13110.place(x=192,y=400)
select_item['tg13120']=0;select_13120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13120'],command=lambda:click_equipment(13120))
select_13120.place(x=192,y=430)
select_item['tg13130']=0;select_13130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13130'],command=lambda:click_equipment(13130))
select_13130.place(x=192,y=460)
select_item['tg13140']=0;select_13140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13140'],command=lambda:click_equipment(13140))
select_13140.place(x=192,y=490)
select_item['tg13150']=0;select_13150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13150'],command=lambda:click_equipment(13150))
select_13150.place(x=192,y=520) # 这里是大自然
## 腰带
select_item['tg14010']=0;select_14010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14010'],command=lambda:click_equipment(14010))
select_14010.place(x=223,y=100)
select_item['tg14020']=0;select_14020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14020'],command=lambda:click_equipment(14020))
select_14020.place(x=223,y=130)
select_item['tg14030']=0;select_14030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14030'],command=lambda:click_equipment(14030))
select_14030.place(x=223,y=160)
select_item['tg14040']=0;select_14040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14040'],command=lambda:click_equipment(14040))
select_14040.place(x=223,y=190)
select_item['tg14050']=0;select_14050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14050'],command=lambda:click_equipment(14050))
select_14050.place(x=223,y=220)
select_item['tg14060']=0;select_14060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14060'],command=lambda:click_equipment(14060))
select_14060.place(x=223,y=250)
select_item['tg14070']=0;select_14070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14070'],command=lambda:click_equipment(14070))
select_14070.place(x=223,y=280)
select_item['tg14080']=0;select_14080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14080'],command=lambda:click_equipment(14080))
select_14080.place(x=223,y=310)
select_item['tg14090']=0;select_14090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14090'],command=lambda:click_equipment(14090))
select_14090.place(x=223,y=340)
select_item['tg14100']=0;select_14100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14100'],command=lambda:click_equipment(14100))
select_14100.place(x=223,y=370)
select_item['tg14110']=0;select_14110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14110'],command=lambda:click_equipment(14110))
select_14110.place(x=223,y=400)
select_item['tg14120']=0;select_14120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14120'],command=lambda:click_equipment(14120))
select_14120.place(x=223,y=430)
select_item['tg14130']=0;select_14130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14130'],command=lambda:click_equipment(14130))
select_14130.place(x=223,y=460)
select_item['tg14140']=0;select_14140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14140'],command=lambda:click_equipment(14140))
select_14140.place(x=223,y=490)
select_item['tg14150']=0;select_14150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14150'],command=lambda:click_equipment(14150))
select_14150.place(x=223,y=520)
## 鞋子
select_item['tg15010']=0;select_15010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15010'],command=lambda:click_equipment(15010))
select_15010.place(x=254,y=100)
select_item['tg15020']=0;select_15020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15020'],command=lambda:click_equipment(15020))
select_15020.place(x=254,y=130)
select_item['tg15030']=0;select_15030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15030'],command=lambda:click_equipment(15030))
select_15030.place(x=254,y=160)
select_item['tg15040']=0;select_15040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15040'],command=lambda:click_equipment(15040))
select_15040.place(x=254,y=190)
select_item['tg15050']=0;select_15050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15050'],command=lambda:click_equipment(15050))
select_15050.place(x=254,y=220)
select_item['tg15060']=0;select_15060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15060'],command=lambda:click_equipment(15060))
select_15060.place(x=254,y=250)
select_item['tg15070']=0;select_15070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15070'],command=lambda:click_equipment(15070))
select_15070.place(x=254,y=280)
select_item['tg15080']=0;select_15080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15080'],command=lambda:click_equipment(15080))
select_15080.place(x=254,y=310)
select_item['tg15090']=0;select_15090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15090'],command=lambda:click_equipment(15090))
select_15090.place(x=254,y=340)
select_item['tg15100']=0;select_15100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15100'],command=lambda:click_equipment(15100))
select_15100.place(x=254,y=370)
select_item['tg15110']=0;select_15110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15110'],command=lambda:click_equipment(15110))
select_15110.place(x=254,y=400)
select_item['tg15120']=0;select_15120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15120'],command=lambda:click_equipment(15120))
select_15120.place(x=254,y=430)
select_item['tg15130']=0;select_15130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15130'],command=lambda:click_equipment(15130))
select_15130.place(x=254,y=460)
select_item['tg15140']=0;select_15140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15140'],command=lambda:click_equipment(15140))
select_15140.place(x=254,y=490)
select_item['tg15150']=0;select_15150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15150'],command=lambda:click_equipment(15150))
select_15150.place(x=254,y=520)
select_item['tg15320']=0;select_15320=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15320'],command=lambda:click_equipment(15320))
select_15320.place(x=492,y=570)
select_item['tg15330']=0;select_15330=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15330'],command=lambda:click_equipment(15330))
select_15330.place(x=492,y=600)
select_item['tg15340']=0;select_15340=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15340'],command=lambda:click_equipment(15340))
select_15340.place(x=492,y=630)
select_item['tg15350']=0;select_15350=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15350'],command=lambda:click_equipment(15350))
select_15350.place(x=492,y=660)
## 首饰-手镯
select_item['tg21160']=0;select_21160=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21160'],command=lambda:click_equipment(21160))
select_21160.place(x=370-12,y=100)
select_item['tg21161']=0;select_21161=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21161'],command=lambda:click_equipment(21161))
select_21161.place(x=370-12+30,y=100)
select_item['tg21170']=0;select_21170=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21170'],command=lambda:click_equipment(21170))
select_21170.place(x=370-12,y=130)
select_item['tg21171']=0;select_21171=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21171'],command=lambda:click_equipment(21171))
select_21171.place(x=370-12+30,y=130)
select_item['tg21180']=0;select_21180=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21180'],command=lambda:click_equipment(21180))
select_21180.place(x=370-12,y=160)
select_item['tg21181']=0;select_21181=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21181'],command=lambda:click_equipment(21181))
select_21181.place(x=370-12+30,y=160)
select_item['tg21190']=0;select_21190=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21190'],command=lambda:click_equipment(21190))
select_21190.place(x=370-12,y=190)
select_item['tg21191']=0;select_21191=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21191'],command=lambda:click_equipment(21191))
select_21191.place(x=370-12+30,y=190)
select_item['tg21240']=0;select_21240=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21240'],command=lambda:click_equipment(21240))
select_21240.place(x=327,y=570)
select_item['tg21241']=0;select_21241=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21241'],command=lambda:click_equipment(21241))
select_21241.place(x=357,y=570)
select_item['tg21250']=0;select_21250=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21250'],command=lambda:click_equipment(21250))
select_21250.place(x=327,y=600)
select_item['tg21251']=0;select_21251=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21251'],command=lambda:click_equipment(21251))
select_21251.place(x=357,y=600)
select_item['tg21260']=0;select_21260=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21260'],command=lambda:click_equipment(21260))
select_21260.place(x=327,y=630)
select_item['tg21261']=0;select_21261=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21261'],command=lambda:click_equipment(21261))
select_21261.place(x=357,y=630)
select_item['tg21270']=0;select_21270=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21270'],command=lambda:click_equipment(21270))
select_21270.place(x=327,y=660)
select_item['tg21271']=0;select_21271=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21271'],command=lambda:click_equipment(21271))
select_21271.place(x=357,y=660)
## 首饰- 项链
select_item['tg22160']=0;select_22160=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22160'],command=lambda:click_equipment(22160))
select_22160.place(x=419,y=100)
select_item['tg22170']=0;select_22170=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22170'],command=lambda:click_equipment(22170))
select_22170.place(x=419,y=130)
select_item['tg22180']=0;select_22180=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22180'],command=lambda:click_equipment(22180))
select_22180.place(x=419,y=160)
select_item['tg22190']=0;select_22190=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22190'],command=lambda:click_equipment(22190))
select_22190.place(x=419,y=190)
select_item['tg22280']=0;select_22280=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22280'],command=lambda:click_equipment(22280))
select_22280.place(x=161,y=570)
select_item['tg22290']=0;select_22290=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22290'],command=lambda:click_equipment(22290))
select_22290.place(x=161,y=600)
select_item['tg22300']=0;select_22300=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22300'],command=lambda:click_equipment(22300))
select_22300.place(x=161,y=630)
select_item['tg22310']=0;select_22310=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22310'],command=lambda:click_equipment(22310))
select_22310.place(x=161,y=660)
## 首饰-戒指
select_item['tg23160']=0;select_23160=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23160'],command=lambda:click_equipment(23160))
select_23160.place(x=450,y=100)
select_item['tg23170']=0;select_23170=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23170'],command=lambda:click_equipment(23170))
select_23170.place(x=450,y=130)
select_item['tg23180']=0;select_23180=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23180'],command=lambda:click_equipment(23180))
select_23180.place(x=450,y=160)
select_item['tg23190']=0;select_23190=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23190'],command=lambda:click_equipment(23190))
select_23190.place(x=450,y=190)
select_item['tg23320']=0;select_23320=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23320'],command=lambda:click_equipment(23320))
select_23320.place(x=523,y=570)
select_item['tg23330']=0;select_23330=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23330'],command=lambda:click_equipment(23330))
select_23330.place(x=523,y=600)
select_item['tg23340']=0;select_23340=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23340'],command=lambda:click_equipment(23340))
select_23340.place(x=523,y=630)
select_item['tg23350']=0;select_23350=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23350'],command=lambda:click_equipment(23350))
select_23350.place(x=523,y=660)
## 特殊-辅助装备 （包括散搭里面的）
select_item['tg31200']=0;select_31200=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31200'],command=lambda:click_equipment(31200))
select_31200.place(x=554,y=100)
select_item['tg31210']=0;select_31210=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31210'],command=lambda:click_equipment(31210))
select_31210.place(x=554,y=130)
select_item['tg31220']=0;select_31220=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31220'],command=lambda:click_equipment(31220))
select_31220.place(x=554,y=160)
select_item['tg31230']=0;select_31230=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31230'],command=lambda:click_equipment(31230))
select_31230.place(x=554,y=190)
select_item['tg31280']=0;select_31280=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31280'],command=lambda:click_equipment(31280))
select_31280.place(x=192,y=570)
select_item['tg31290']=0;select_31290=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31290'],command=lambda:click_equipment(31290))
select_31290.place(x=192,y=600)
select_item['tg31300']=0;select_31300=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31300'],command=lambda:click_equipment(31300))
select_31300.place(x=192,y=630)
select_item['tg31310']=0;select_31310=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31310'],command=lambda:click_equipment(31310))
select_31310.place(x=192,y=660)
## 特殊- 魔法石
select_item['tg32200']=0;select_32200=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32200'],command=lambda:click_equipment(32200))
select_32200.place(x=585,y=100)
select_item['tg32210']=0;select_32210=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32210'],command=lambda:click_equipment(32210))
select_32210.place(x=585,y=130)
select_item['tg32220']=0;select_32220=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32220'],command=lambda:click_equipment(32220))
select_32220.place(x=585,y=160)
select_item['tg32230']=0;select_32230=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32230'],command=lambda:click_equipment(32230))
select_32230.place(x=585,y=190)
select_item['tg32240']=0;select_32240=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32240'],command=lambda:click_equipment(32240))
select_32240.place(x=388,y=570)
select_item['tg32250']=0;select_32250=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32250'],command=lambda:click_equipment(32250))
select_32250.place(x=388,y=600)
select_item['tg32260']=0;select_32260=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32260'],command=lambda:click_equipment(32260))
select_32260.place(x=388,y=630)
select_item['tg32270']=0;select_32270=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32270'],command=lambda:click_equipment(32270))
select_32270.place(x=388,y=660)
## 特殊 - 耳环 （包括神话）
select_item['tg33200']=0;select_33200=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33200'],command=lambda:click_equipment(33200))
select_33200.place(x=616,y=100)
select_item['tg33201']=0;select_33201=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33201'],command=lambda:click_equipment(33201))
select_33201.place(x=646,y=100)
select_item['tg33210']=0;select_33210=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33210'],command=lambda:click_equipment(33210))
select_33210.place(x=616,y=130)
select_item['tg33211']=0;select_33211=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33211'],command=lambda:click_equipment(33211))
select_33211.place(x=646,y=130)
select_item['tg33220']=0;select_33220=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33220'],command=lambda:click_equipment(33220))
select_33220.place(x=616,y=160)
select_item['tg33221']=0;select_33221=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33221'],command=lambda:click_equipment(33221))
select_33221.place(x=646,y=160)
select_item['tg33230']=0;select_33230=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33230'],command=lambda:click_equipment(33230))
select_33230.place(x=616,y=190)
select_item['tg33231']=0;select_33231=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33231'],command=lambda:click_equipment(33231))
select_33231.place(x=646,y=190)
select_item['tg33320']=0;select_33320=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33320'],command=lambda:click_equipment(33320))
select_33320.place(x=554,y=570)
select_item['tg33321']=0;select_33321=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33321'],command=lambda:click_equipment(33321))
select_33321.place(x=584,y=570)
select_item['tg33330']=0;select_33330=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33330'],command=lambda:click_equipment(33330))
select_33330.place(x=554,y=600)
select_item['tg33331']=0;select_33331=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33331'],command=lambda:click_equipment(33331))
select_33331.place(x=584,y=600)
select_item['tg33340']=0;select_33340=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33340'],command=lambda:click_equipment(33340))
select_33340.place(x=554,y=630)
select_item['tg33341']=0;select_33341=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33341'],command=lambda:click_equipment(33341))
select_33341.place(x=584,y=630)
select_item['tg33350']=0;select_33350=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33350'],command=lambda:click_equipment(33350))
select_33350.place(x=554,y=660)
select_item['tg33351']=0;select_33351=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33351'],command=lambda:click_equipment(33351))
select_33351.place(x=584,y=660)


def donate():
    webbrowser.open('https://twip.kr/dawnclass16')
donate_image=PhotoImage(file='ext_img/donate.png')
donate_bt=tkinter.Button(self,image=donate_image, command=donate,borderwidth=0,bg=dark_main,activebackground=dark_main)
donate_bt.place(x=625,y=550-28)
def dunfaoff():
    webbrowser.open('https://space.bilibili.com/4952736')
dunfaoff_image=PhotoImage(file='ext_img/dunfaoff.png')
dunfaoff_url=tkinter.Button(self,image=dunfaoff_image, command=dunfaoff,borderwidth=0,bg=dark_main,activebackground=dark_main)
dunfaoff_url.place(x=535,y=410)

def blog():
    webbrowser.open('https://blog.naver.com/dawnclass16/221837654941')
blog_image=PhotoImage(file='ext_img/blog.png')
blog_url=tkinter.Button(self,image=blog_image, command=blog,borderwidth=0,bg=dark_main,activebackground=dark_main)
blog_url.place(x=615,y=410)
    
def hamjung():
    tkinter.messagebox.showinfo("제작자 크레딧","총제작자=Dawnclass(새벽반)\n이미지/그래픽=경철부동산\n직업/버퍼DB=대략볼록할철\n서버제공=던파오프\n기타조언=히든 도비 4,5,6호\n\n오류 제보는 블로그 덧글이나 던조 쪽지로")
maker_image=PhotoImage(file='ext_img/maker.png')
maker=tkinter.Button(self,image=maker_image, command=hamjung,borderwidth=0,bg=dark_main,activebackground=dark_main)
version=tkinter.Label(self,text='V '+str(now_version)+'\n'+ver_time,font=guide_font,fg="white",bg=dark_main)

maker.place(x=625,y=590)
version.place(x=630,y=650)

try:
    html_version=requests.get('https://dunfaoff.com/DawnClass.df')
    soup_version = BeautifulSoup(html_version.content, 'html.parser')
    now_version_num=int(now_version[0]+now_version[2]+now_version[4])
    net_version=str(soup_version)[2:]
    net_version_num=int(net_version[0]+net_version[2]+net_version[4])
    if now_version_num < now_version_num:
        ask_update=tkinter.messagebox.askquestion('업데이트',"最新版本이 존재합니다. 이동하시겠습니까?")
        if ask_update == 'yes':
            webbrowser.open('https://drive.google.com/open?id=1p8ZdzW_NzGKHHOtfPTuZSr1YgSEVtYCj')
    else:
        print("最新版本")
except:
    print("更新检查失败（网络错误）")

if __name__ == "__main__":
    update_thread()
self.mainloop()


self.quit()

